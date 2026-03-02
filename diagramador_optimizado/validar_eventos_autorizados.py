"""
Script para validar que no se creen eventos no autorizados:
- Desplazamientos no habilitados
- Vacíos no habilitados
- Paradas fuera de rango permitido
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional

def cargar_configuracion() -> Dict[str, Any]:
    """Carga la configuración desde configuracion.json"""
    config_path = Path(__file__).parent / "configuracion.json"
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def verificar_vacio_habilitado(config: Dict[str, Any], origen: str, destino: str) -> bool:
    """Verifica si un vacío está habilitado en la configuración"""
    vacios = config.get("vacios", {})
    clave = f"{origen}_{destino}"
    entrada = vacios.get(clave, {})
    
    if isinstance(entrada, dict):
        # Por defecto, si no se especifica habilitado, se considera True
        return entrada.get("habilitado", True)
    
    # Si no existe la entrada, no está habilitado
    return False

def verificar_desplazamiento_habilitado(config: Dict[str, Any], origen: str, destino: str) -> bool:
    """Verifica si un desplazamiento está habilitado en la configuración"""
    desplazamientos = config.get("desplazamientos", {})
    clave = f"{origen}_{destino}"
    entrada = desplazamientos.get(clave, {})
    
    if isinstance(entrada, dict):
        # Por defecto, si no se especifica habilitado, se considera False para desplazamientos
        return entrada.get("habilitado", False)
    
    # Si no existe la entrada, no está habilitado
    return False

def verificar_parada_valida(config: Dict[str, Any], nodo: str, duracion: int) -> bool:
    """Verifica si una parada está dentro del rango permitido"""
    paradas = config.get("paradas", {})
    regla = paradas.get(nodo.upper(), {})
    
    if not regla:
        # Si no hay regla específica, permitir cualquier duración
        return True
    
    parada_min = regla.get("min", 0)
    parada_max = regla.get("max", 1440)
    
    return parada_min <= duracion <= parada_max

def analizar_eventos_excel(archivo_excel: Optional[str] = None) -> Dict[str, List[str]]:
    """
    Analiza los eventos generados para detectar eventos no autorizados.
    Si no se proporciona archivo_excel, intenta leer el archivo de salida por defecto.
    """
    import openpyxl
    
    if archivo_excel is None:
        archivo_excel = Path(__file__).parent / "resultado_diagramacion.xlsx"
    
    if not Path(archivo_excel).exists():
        return {"error": [f"Archivo no encontrado: {archivo_excel}"]}
    
    config = cargar_configuracion()
    errores = {
        "vacios_no_habilitados": [],
        "desplazamientos_no_habilitados": [],
        "paradas_invalidas": [],
    }
    
    wb = openpyxl.load_workbook(archivo_excel)
    
    # Buscar la hoja de eventos (puede llamarse "Eventos" o similar)
    nombre_hoja_eventos = None
    for nombre_hoja in wb.sheetnames:
        if "evento" in nombre_hoja.lower():
            nombre_hoja_eventos = nombre_hoja
            break
    
    if not nombre_hoja_eventos:
        return {"error": [f"Hoja de eventos no encontrada. Hojas disponibles: {', '.join(wb.sheetnames)}"]}
    
    ws = wb[nombre_hoja_eventos]
    
    # Leer encabezados
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        headers[cell.value] = col_idx
    
    # Leer eventos
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        evento = {}
        for header, col_idx in headers.items():
            if col_idx <= len(row):
                evento[header] = row[col_idx - 1].value
        
        tipo_evento = str(evento.get("Evento", "")).strip().upper()
        origen = str(evento.get("Origen", "")).strip()
        destino = str(evento.get("Destino", "")).strip()
        
        # Convertir inicio y fin a números (pueden venir como strings o números)
        inicio_val = evento.get("Inicio", 0)
        fin_val = evento.get("Fin", 0)
        try:
            inicio = float(inicio_val) if inicio_val else 0
            fin = float(fin_val) if fin_val else 0
        except (ValueError, TypeError):
            inicio = 0
            fin = 0
        
        duracion = int(fin - inicio) if fin and inicio else 0
        
        # Validar Vacíos
        if tipo_evento == "VACIO":
            if not verificar_vacio_habilitado(config, origen, destino):
                errores["vacios_no_habilitados"].append(
                    f"Fila {row_idx}: Vacio {origen} -> {destino} no está habilitado"
                )
        
        # Validar Desplazamientos
        elif tipo_evento == "DESPLAZAMIENTO":
            if not verificar_desplazamiento_habilitado(config, origen, destino):
                errores["desplazamientos_no_habilitados"].append(
                    f"Fila {row_idx}: Desplazamiento {origen} -> {destino} no está habilitado"
                )
        
        # Validar Paradas
        elif tipo_evento == "PARADA":
            if not verificar_parada_valida(config, origen, duracion):
                regla = config.get("paradas", {}).get(origen.upper(), {})
                parada_min = regla.get("min", 0) if regla else 0
                parada_max = regla.get("max", 1440) if regla else 1440
                errores["paradas_invalidas"].append(
                    f"Fila {row_idx}: Parada en {origen} con duración {duracion} min "
                    f"fuera de rango permitido ({parada_min}-{parada_max} min)"
                )
    
    return errores

def main():
    """Función principal"""
    print("=" * 80)
    print("VALIDACIÓN DE EVENTOS AUTORIZADOS")
    print("=" * 80)
    print()
    
    errores = analizar_eventos_excel()
    
    if "error" in errores:
        print(f"[ERROR] {errores['error'][0]}")
        return
    
    total_errores = (
        len(errores["vacios_no_habilitados"]) +
        len(errores["desplazamientos_no_habilitados"]) +
        len(errores["paradas_invalidas"])
    )
    
    if total_errores == 0:
        print("[OK] Todos los eventos están autorizados según la configuración")
        print()
        return
    
    print(f"[ERROR] Se encontraron {total_errores} eventos no autorizados:")
    print()
    
    if errores["vacios_no_habilitados"]:
        print(f"Vacios no habilitados ({len(errores['vacios_no_habilitados'])}):")
        for error in errores["vacios_no_habilitados"][:20]:
            print(f"  - {error}")
        if len(errores["vacios_no_habilitados"]) > 20:
            print(f"  ... y {len(errores['vacios_no_habilitados']) - 20} más")
        print()
    
    if errores["desplazamientos_no_habilitados"]:
        print(f"Desplazamientos no habilitados ({len(errores['desplazamientos_no_habilitados'])}):")
        for error in errores["desplazamientos_no_habilitados"][:20]:
            print(f"  - {error}")
        if len(errores["desplazamientos_no_habilitados"]) > 20:
            print(f"  ... y {len(errores['desplazamientos_no_habilitados']) - 20} más")
        print()
    
    if errores["paradas_invalidas"]:
        print(f"Paradas inválidas ({len(errores['paradas_invalidas'])}):")
        for error in errores["paradas_invalidas"][:20]:
            print(f"  - {error}")
        if len(errores["paradas_invalidas"]) > 20:
            print(f"  ... y {len(errores['paradas_invalidas']) - 20} más")
        print()

if __name__ == "__main__":
    main()
