"""
Script para verificar continuidad de eventos de conductores en resultado_diagramacion.xlsx.
Regla transporte: fin(N) = inicio(N+1) o gap válido (layover en terminal).
- GAP válido: Parada/Vacio -> Comercial/Parada en mismo nodo, diff 1-45 min (layover).
- diff < 0: ERROR (solapamiento u orden incorrecto).
- diff > 45 min: ERROR (hueco excesivo).
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Necesita openpyxl: pip install openpyxl")
    sys.exit(1)


def hhmm_a_minutos(val) -> int:
    """Convierte HH:MM o valor numérico a minutos."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if not s:
        return 0
    try:
        partes = s.split(":")
        h = int(partes[0])
        m = int(partes[1]) if len(partes) > 1 else 0
        return h * 60 + m
    except Exception:
        return 0


def main():
    import sys as _sys
    base = Path(__file__).resolve().parent
    excel_name = _sys.argv[1] if len(_sys.argv) > 1 else "resultado_diagramacion.xlsx"
    excel_path = base / excel_name
    if not excel_path.exists():
        print(f"No encontrado: {excel_path}")
        return 1

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "EventosCompletos" not in wb.sheetnames:
        print("Hoja EventosCompletos no encontrada")
        return 1

    ws = wb["EventosCompletos"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Columnas: Evento, Bus, Conductor, Inicio, Fin, Duración, Origen, Destino, ...
    idx_evento = 0
    idx_conductor = 2
    idx_inicio = 3
    idx_fin = 4
    idx_origen = 6
    idx_destino = 7

    # Agrupar por (conductor, bus) - un conductor puede tener varios buses/segmentos
    eventos_por_clave: dict[tuple, list] = {}
    for row in rows:
        if not row or len(row) <= max(idx_evento, idx_conductor, idx_fin):
            continue
        ev = str(row[idx_evento] or "").strip()
        cond = row[idx_conductor]
        if cond is None or cond == "":
            continue
        try:
            cid = int(cond)
        except (TypeError, ValueError):
            continue
        bus_val = row[1] if len(row) > 1 else ""
        bus_key = bus_val if bus_val not in (None, "") else "sin_bus"
        inicio = hhmm_a_minutos(row[idx_inicio])
        fin = hhmm_a_minutos(row[idx_fin])
        origen = str(row[idx_origen] or "")
        destino = str(row[idx_destino] or "")

        clave = (cid, bus_key)
        if clave not in eventos_por_clave:
            eventos_por_clave[clave] = []
        eventos_por_clave[clave].append({
            "evento": ev,
            "inicio": inicio,
            "fin": fin,
            "origen": origen,
            "destino": destino,
            "bus": bus_val,
        })
    eventos_por_conductor = {c: [] for c in set(k[0] for k in eventos_por_clave)}
    for (c, b), evs in eventos_por_clave.items():
        for e in evs:
            eventos_por_conductor[c].append(e)

    wb.close()

    GAP_LAYOVER_MAX = 45  # min: layover válido en terminal (Parada/Vacio -> siguiente evento)
    errores = []
    warnings = []
    segmentos_ok = 0
    for (cid, bus_key), evs in sorted(eventos_por_clave.items(), key=lambda x: (x[0][0], str(x[0][1]))):
        if bus_key == "sin_bus":
            continue
        evs_ord = sorted(evs, key=lambda x: (x["inicio"], x["fin"]))
        tiene_error = False
        for i in range(1, len(evs_ord)):
            ant = evs_ord[i - 1]
            act = evs_ord[i]
            fin_ant = ant["fin"]
            ini_act = act["inicio"]
            diff = ini_act - fin_ant
            if diff == 0:
                continue
            dest_ant = (ant.get("destino") or "").strip().upper()
            orig_act = (act.get("origen") or "").strip().upper()
            mismo_nodo = dest_ant and orig_act and (dest_ant == orig_act or dest_ant in orig_act or orig_act in dest_ant)
            # diff < 0: solapamiento u orden incorrecto -> ERROR
            if diff < 0:
                tiene_error = True
                errores.append({
                    "conductor": cid,
                    "bus": bus_key,
                    "evento_ant": ant["evento"],
                    "evento_act": act["evento"],
                    "fin_ant": fin_ant,
                    "ini_act": ini_act,
                    "diff": diff,
                    "horas_ant": f"{fin_ant//60:02d}:{fin_ant%60:02d}",
                    "horas_act": f"{ini_act//60:02d}:{ini_act%60:02d}",
                    "tipo": "solapamiento",
                })
            elif diff <= GAP_LAYOVER_MAX and mismo_nodo:
                # Layover válido: Parada/Vacio termina, siguiente empieza en mismo nodo
                continue
            elif diff > GAP_LAYOVER_MAX:
                tiene_error = True
                errores.append({
                    "conductor": cid,
                    "bus": bus_key,
                    "evento_ant": ant["evento"],
                    "evento_act": act["evento"],
                    "fin_ant": fin_ant,
                    "ini_act": ini_act,
                    "diff": diff,
                    "horas_ant": f"{fin_ant//60:02d}:{fin_ant%60:02d}",
                    "horas_act": f"{ini_act//60:02d}:{ini_act%60:02d}",
                    "tipo": "hueco_excesivo",
                })
            else:
                warnings.append({
                    "conductor": cid,
                    "bus": bus_key,
                    "diff": diff,
                    "horas_ant": f"{fin_ant//60:02d}:{fin_ant%60:02d}",
                    "horas_act": f"{ini_act//60:02d}:{ini_act%60:02d}",
                })
        if not tiene_error:
            segmentos_ok += 1

    # Resumen
    print("=" * 70)
    print("VERIFICACIÓN DE CONTINUIDAD - EventosCompletos (por conductor+bus)")
    print("=" * 70)
    print(f"Segmentos (cond+bus) analizados: {len(eventos_por_clave)}")
    print(f"Segmentos con secuencia OK: {segmentos_ok}")
    print(f"Total huecos detectados: {len(errores)}")
    print()

    if errores:
        err_solap = [e for e in errores if e.get("tipo") == "solapamiento"]
        err_hueco = [e for e in errores if e.get("tipo") == "hueco_excesivo"]
        print("Errores críticos (solapamiento / orden incorrecto):", len(err_solap))
        for e in err_solap[:15]:
            bus_str = f" bus={e.get('bus', '')}" if e.get('bus') else ""
            print(f"  Cond {e['conductor']}{bus_str}: {e['evento_ant']} termina {e['horas_ant']} -> "
                  f"{e['evento_act']} empieza {e['horas_act']} (diff={e['diff']} min)")
        if len(err_solap) > 15:
            print(f"  ... y {len(err_solap) - 15} más")
        print("Huecos > 45 min:", len(err_hueco))
        for e in err_hueco[:5]:
            bus_str = f" bus={e.get('bus', '')}" if e.get('bus') else ""
            print(f"  Cond {e['conductor']}{bus_str}: diff={e['diff']} min")
        if len(err_hueco) > 5:
            print(f"  ... y {len(err_hueco) - 5} más")
    else:
        print("[OK] Todos los segmentos tienen secuencia continua (fin(N)=inicio(N+1)).")

    # Verificación específica Conductor 2 (caso Vacio LOS TILOS)
    if 2 in eventos_por_conductor:
        evs2 = sorted(eventos_por_conductor[2], key=lambda x: (x["inicio"], x["fin"]))
        print()
        print("Conductor 2 - Secuencia de eventos:")
        for ev in evs2[:12]:
            h_i = f"{ev['inicio']//60:02d}:{ev['inicio']%60:02d}"
            h_f = f"{ev['fin']//60:02d}:{ev['fin']%60:02d}"
            print(f"  {ev['evento']:15} {h_i}-{h_f}  {ev['origen']} -> {ev['destino']}  (bus={ev['bus']})")
        if len(evs2) > 12:
            print(f"  ... ({len(evs2)} eventos total)")

    print("=" * 70)
    return 0 if not errores else 1


if __name__ == "__main__":
    sys.exit(main())
