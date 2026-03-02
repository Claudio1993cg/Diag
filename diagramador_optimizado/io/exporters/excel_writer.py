# -*- coding: utf-8 -*-
"""
Exportación a Excel - Lógica de Origen con corte Fase 2 y puntos de relevo.

- Genera eventos de bus: Vacio y Parada desde bloques; Comercial solo desde insumo de carga.
- Usa la misma lógica de corte y puntos de relevo que Fase 2: cortes solo en depósito
  o nodos con desplazamiento bidireccional al depósito; fin de turno en relay = conductor
  se baja al terminar el último viaje (no se pinta después en ese bus).
- Asigna conductores por mapeo minuto-a-minuto respetando esos rangos.
- InS, FnS y Desplazamiento coherentes con relevos (Desplazamiento solo cuando no hay vacío).
"""
from __future__ import annotations

import collections
import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from diagramador_optimizado.core.domain.logistica import GestorDeLogistica
from diagramador_optimizado.core.engines.fase2_conductores import (
    _es_deposito,
    _es_relevo_valido,
    _calcular_fin_turno,
)


def _origen_es_deposito_estricto(origen: str, gestor: GestorDeLogistica) -> bool:
    """True solo si origen es exactamente el depósito (evita que 'PIE ANDINO' se confunda con 'Deposito Pie Andino')."""
    o = (origen or "").strip().upper()
    if not o:
        return False
    nombres_dep = gestor._nombres_depositos() if hasattr(gestor, "_nombres_depositos") else None
    for d in (nombres_dep or [gestor.deposito_base]):
        if (d or "").strip().upper() == o:
            return True
    return False
from diagramador_optimizado.utils.time_utils import formatear_hora, formatear_hora_deltatime, _to_minutes


def _tiempo_a_minutos(val: Any) -> int:
    """Convierte HH:MM o número a minutos."""
    return _to_minutes(val)


def _obtener_rangos_parada(gestor: Optional[GestorDeLogistica], nombre_nodo: str) -> Tuple[int, int]:
    """Parada min/max para un nodo."""
    if not nombre_nodo or not gestor:
        return 0, 24 * 60
    regla = getattr(gestor, "paradas_dict", None) or {}
    regla = regla.get(str(nombre_nodo).strip().upper(), {}) if isinstance(regla, dict) else {}
    return regla.get("min", 0), regla.get("max", 24 * 60)


def _generar_eventos_de_bus_fase1(
    gestor: GestorDeLogistica,
    bloques_bus: List[List[Dict[str, Any]]],
) -> Tuple[List[Dict[str, Any]], float]:
    """
    Genera solo eventos Vacio y Parada desde bloques.
    Los comerciales se generan únicamente desde el insumo de carga (viajes_comerciales).
    Devuelve (eventos_fase1_buses, km_vacio_total).
    """
    deposito = gestor.deposito_base
    eventos_fase1_buses: List[Dict[str, Any]] = []
    km_vacio_total = 0.0
    # Los únicos comerciales son los del insumo de carga; desde bloques solo se generan Vacio y Parada.

    for bus_id_idx, bloque in enumerate(bloques_bus):
        bus_id = bus_id_idx + 1
        if not bloque:
            continue

        primer_viaje = bloque[0]
        parada_min_origen, parada_max_origen = _obtener_rangos_parada(gestor, primer_viaje.get("origen", ""))

        # Vacío inicial (depósito -> primer origen)
        if not _es_deposito(primer_viaje.get("origen", ""), deposito):
            tiempo_vacio_ini, km_vacio_ini = gestor.buscar_tiempo_vacio(
                deposito, primer_viaje["origen"], primer_viaje["inicio"]
            )
            if tiempo_vacio_ini is None or tiempo_vacio_ini <= 0:
                tiempo_vacio_ini = 0
                km_vacio_ini = 0
            fin_vacio_ini = primer_viaje["inicio"] - parada_min_origen
            inicio_vacio_ini = fin_vacio_ini - tiempo_vacio_ini
            if inicio_vacio_ini < 0:
                inicio_vacio_ini = 0
                fin_vacio_ini = inicio_vacio_ini + tiempo_vacio_ini

            eventos_fase1_buses.append({
                "evento": "Vacio",
                "bus": bus_id,
                "conductor": "",
                "inicio": formatear_hora_deltatime(inicio_vacio_ini),
                "fin": formatear_hora_deltatime(fin_vacio_ini),
                "duracion": formatear_hora_deltatime(tiempo_vacio_ini),
                "origen": deposito,
                "destino": primer_viaje["origen"],
                "linea": "",
                "kilometros": km_vacio_ini or 0,
                "desc": f"Vacío {deposito}->{primer_viaje['origen']}",
            })
            km_vacio_total += km_vacio_ini or 0

            tiempo_parada_ini = primer_viaje["inicio"] - (fin_vacio_ini if tiempo_vacio_ini else primer_viaje["inicio"])
            if tiempo_parada_ini > 0:
                eventos_fase1_buses.append({
                    "evento": "Parada",
                    "bus": bus_id,
                    "conductor": "",
                    "inicio": formatear_hora_deltatime(fin_vacio_ini),
                    "fin": formatear_hora_deltatime(primer_viaje["inicio"]),
                    "duracion": formatear_hora_deltatime(tiempo_parada_ini),
                    "origen": primer_viaje["origen"],
                    "destino": primer_viaje["origen"],
                    "linea": "",
                    "kilometros": 0,
                    "desc": f"Parada en {primer_viaje['origen']}",
                })
        else:
            hora_inicio_parada = max(primer_viaje["inicio"] - parada_min_origen, 0)
            tiempo_parada_ini = primer_viaje["inicio"] - hora_inicio_parada
            if tiempo_parada_ini > 0:
                eventos_fase1_buses.append({
                    "evento": "Parada",
                    "bus": bus_id,
                    "conductor": "",
                    "inicio": formatear_hora_deltatime(hora_inicio_parada),
                    "fin": formatear_hora_deltatime(primer_viaje["inicio"]),
                    "duracion": formatear_hora_deltatime(tiempo_parada_ini),
                    "origen": primer_viaje["origen"],
                    "destino": primer_viaje["origen"],
                    "linea": "",
                    "kilometros": 0,
                    "desc": f"Parada en {primer_viaje['origen']}",
                })

        for indice_viaje, viaje in enumerate(bloque):
            if indice_viaje >= len(bloque) - 1:
                continue
            siguiente_viaje = bloque[indice_viaje + 1]
            if viaje["destino"] != siguiente_viaje["origen"]:
                tiempo_vacio_inter, km_vacio_inter = gestor.buscar_tiempo_vacio(
                    viaje["destino"], siguiente_viaje["origen"], viaje["fin"]
                )
                if tiempo_vacio_inter is None or tiempo_vacio_inter <= 0:
                    tiempo_vacio_inter, km_vacio_inter = 0, 0
                inicio_vacio = viaje["fin"]
                fin_vacio = inicio_vacio + tiempo_vacio_inter
                eventos_fase1_buses.append({
                    "evento": "Vacio",
                    "bus": bus_id,
                    "conductor": "",
                    "inicio": formatear_hora_deltatime(inicio_vacio),
                    "fin": formatear_hora_deltatime(fin_vacio),
                    "duracion": formatear_hora_deltatime(tiempo_vacio_inter),
                    "origen": viaje["destino"],
                    "destino": siguiente_viaje["origen"],
                    "linea": "",
                    "kilometros": km_vacio_inter or 0,
                    "desc": f"Vacío {viaje['destino']}->{siguiente_viaje['origen']}",
                })
                km_vacio_total += km_vacio_inter or 0
                parada_min_dest, parada_max_dest = _obtener_rangos_parada(gestor, siguiente_viaje["origen"])
                tiempo_parada = siguiente_viaje["inicio"] - fin_vacio
                if tiempo_parada > 0:
                    eventos_fase1_buses.append({
                        "evento": "Parada",
                        "bus": bus_id,
                        "conductor": "",
                        "inicio": formatear_hora_deltatime(fin_vacio),
                        "fin": formatear_hora_deltatime(siguiente_viaje["inicio"]),
                        "duracion": formatear_hora_deltatime(tiempo_parada),
                        "origen": siguiente_viaje["origen"],
                        "destino": siguiente_viaje["origen"],
                        "linea": "",
                        "kilometros": 0,
                        "desc": f"Parada en {siguiente_viaje['origen']}",
                    })
            else:
                parada_min_dest, parada_max_dest = _obtener_rangos_parada(gestor, viaje["destino"])
                tiempo_parada = siguiente_viaje["inicio"] - viaje["fin"]
                if tiempo_parada > 0:
                    eventos_fase1_buses.append({
                        "evento": "Parada",
                        "bus": bus_id,
                        "conductor": "",
                        "inicio": formatear_hora_deltatime(viaje["fin"]),
                        "fin": formatear_hora_deltatime(siguiente_viaje["inicio"]),
                        "duracion": formatear_hora_deltatime(tiempo_parada),
                        "origen": viaje["destino"],
                        "destino": viaje["destino"],
                        "linea": "",
                        "kilometros": 0,
                        "desc": f"Parada en {viaje['destino']}",
                    })

        ultimo_viaje = bloque[-1]
        if not _es_deposito(ultimo_viaje.get("destino", ""), deposito):
            tiempo_vacio_fin, km_vacio_fin = gestor.buscar_tiempo_vacio(
                ultimo_viaje["destino"], deposito, ultimo_viaje["fin"]
            )
            if tiempo_vacio_fin is None or tiempo_vacio_fin <= 0:
                tiempo_vacio_fin, km_vacio_fin = 0, 0
            inicio_vacio_fin = ultimo_viaje["fin"]
            fin_vacio_fin = inicio_vacio_fin + tiempo_vacio_fin
            eventos_fase1_buses.append({
                "evento": "Vacio",
                "bus": bus_id,
                "conductor": "",
                "inicio": formatear_hora_deltatime(inicio_vacio_fin),
                "fin": formatear_hora_deltatime(fin_vacio_fin),
                "duracion": formatear_hora_deltatime(tiempo_vacio_fin),
                "origen": ultimo_viaje["destino"],
                "destino": deposito,
                "linea": "",
                "kilometros": km_vacio_fin or 0,
                "desc": f"Vacío {ultimo_viaje['destino']}->{deposito}",
            })
            km_vacio_total += km_vacio_fin or 0

    eventos_fase1_buses.sort(key=lambda e: (_tiempo_a_minutos(e.get("inicio", 0)), e.get("bus", 0)))
    return eventos_fase1_buses, round(km_vacio_total, 1)


def _construir_mapa_viajes(
    viajes_comerciales: List[Dict[str, Any]],
    metadata_tareas: Dict[Any, Dict[str, Any]],
) -> Dict[Any, Dict[str, Any]]:
    """Mapa id_tarea -> viaje (id y _tmp_id)."""
    mapa: Dict[Any, Dict[str, Any]] = {}
    for v in viajes_comerciales or []:
        for key in (v.get("id"), v.get("_tmp_id")):
            if key is not None:
                mapa[key] = v
                mapa[str(key)] = v
    for tid, meta in (metadata_tareas or {}).items():
        v = meta.get("viaje") if isinstance(meta, dict) else None
        if v and tid not in mapa:
            mapa[tid] = v
            mapa[str(tid)] = v
    return mapa


def _construir_conductor_por_bus_tiempo(
    gestor: GestorDeLogistica,
    turnos_seleccionados: List[Dict[str, Any]],
    metadata_tareas: Dict[Any, Dict[str, Any]],
    mapa_viajes: Dict[Any, Dict[str, Any]],
) -> Dict[Tuple[int, int], int]:
    """
    (bus_id, minuto) -> conductor_id (1-based).
    Respeta la lógica de corte Fase 2 y puntos de relevo:
    - Vacío inicial: solo se pinta si existe vacío depot→origen (conductor sube en depot).
    - Fin en relay: el conductor se baja al terminar el último viaje, no se pinta después en ese bus.
    - Fin en depot: el conductor sigue en el bus hasta fin del último viaje (no pintamos vacío vuelta al depot
      para no solapar con posible siguiente conductor; el rango comercial+paradas ya está cubierto).
    """
    deposito = gestor.deposito_base
    conductor_por_bus_tiempo: Dict[Tuple[int, int], int] = {}
    for id_conductor, turno in enumerate(turnos_seleccionados or [], start=1):
        tareas = turno.get("tareas_con_bus", [])
        if not tareas:
            continue
        for idx, (id_tarea, indice_bus) in enumerate(tareas):
            viaje = mapa_viajes.get(id_tarea) or mapa_viajes.get(str(id_tarea))
            if not viaje:
                continue
            bus_id = indice_bus + 1
            for minuto in range(viaje["inicio"], viaje["fin"]):
                conductor_por_bus_tiempo[(bus_id, minuto)] = id_conductor
            if idx < len(tareas) - 1:
                prox_id, prox_bus_idx = tareas[idx + 1]
                if prox_bus_idx == indice_bus:
                    viaje_sig = mapa_viajes.get(prox_id) or mapa_viajes.get(str(prox_id))
                    if viaje_sig and viaje_sig["inicio"] > viaje["fin"]:
                        for minuto in range(viaje["fin"], viaje_sig["inicio"]):
                            conductor_por_bus_tiempo[(bus_id, minuto)] = id_conductor
        primer_id, primer_bus_idx = tareas[0]
        primer_viaje = mapa_viajes.get(primer_id) or mapa_viajes.get(str(primer_id))
        if primer_viaje:
            bus_inicial = primer_bus_idx + 1
            origen_primero = (primer_viaje.get("origen") or "").strip()
            inicio_min_viaje = primer_viaje["inicio"]
            # Regla: vacío inicial solo se traza si existe vacío depot→origen.
            # Si el primer viaje parte del depósito, NO pintamos minutos previos
            # (evita tener InS -> Parada como primer evento del conductor).
            if not _es_deposito(origen_primero, deposito):
                t_vacio_ini, _ = gestor.buscar_tiempo_vacio(deposito, origen_primero, inicio_min_viaje)
                if t_vacio_ini and t_vacio_ini > 0:
                    desde = max(0, inicio_min_viaje - t_vacio_ini)
                    for minuto in range(desde, inicio_min_viaje + 1):
                        conductor_por_bus_tiempo[(bus_inicial, minuto)] = id_conductor
        # Corte Fase 2 / relevo: no pintar después del fin del último viaje (conductor se baja en depot o relay).
    return conductor_por_bus_tiempo


def _generar_eventos_completos(
    gestor: GestorDeLogistica,
    bloques_bus: List[List[Dict[str, Any]]],
    turnos_seleccionados: List[Dict[str, Any]],
    viajes_comerciales: List[Dict[str, Any]],
    metadata_tareas: Dict[Any, Dict[str, Any]],
    eventos_bus: Optional[List[List[Dict[str, Any]]]] = None,
) -> List[Dict[str, Any]]:
    """
    Genera todos los eventos (Fase 1 + asignación conductor + InS, FnS, Desplazamiento).
    Retorna lista plana de eventos para EventosCompletos.
    """
    deposito = gestor.deposito_base
    tiempo_toma = gestor.tiempo_toma

    # 1) Eventos de buses de Fase 1 (Vacio y Parada)
    #    Si recibimos eventos_bus desde el motor, reutilizamos exactamente esos
    #    (para que los tiempos de parada coincidan 1:1 con la Fase 1).
    if eventos_bus:
        eventos_fase1_buses: List[Dict[str, Any]] = []
        for bus_idx, lista in enumerate(eventos_bus):
            bus_id = bus_idx + 1
            for ev in lista or []:
                tipo_raw = (str(ev.get("evento", "")) or "").strip()
                if tipo_raw.upper() not in ("VACIO", "PARADA", "RECARGA"):
                    continue
                inicio_min = int(ev.get("inicio", 0) or 0)
                fin_min = int(ev.get("fin", 0) or 0)
                dur_min = max(0, fin_min - inicio_min)
                ev_copia = {
                    "evento": tipo_raw,
                    "bus": bus_id,
                    "conductor": "",
                    "inicio": formatear_hora_deltatime(inicio_min),
                    "fin": formatear_hora_deltatime(fin_min),
                    "duracion": formatear_hora_deltatime(dur_min),
                    "origen": ev.get("origen", ""),
                    "destino": ev.get("destino", ""),
                    "linea": ev.get("linea", ""),
                    "kilometros": ev.get("kilometros", 0),
                    "desc": ev.get("desc", ""),
                    "tipo_bus": ev.get("tipo_bus", ""),
                }
                if tipo_raw == "Recarga":
                    ev_copia["porcentaje_bateria"] = ev.get("porcentaje_bateria", "")
                    ev_copia["posicion_recarga"] = ev.get("posicion_recarga", "")
                eventos_fase1_buses.append(ev_copia)
    else:
        eventos_fase1_buses, _ = _generar_eventos_de_bus_fase1(gestor, bloques_bus)
    mapa_viajes = _construir_mapa_viajes(viajes_comerciales, metadata_tareas)
    conductor_por_bus_tiempo = _construir_conductor_por_bus_tiempo(
        gestor, turnos_seleccionados, metadata_tareas, mapa_viajes
    )

    eventos_por_bus: Dict[int, List[Dict]] = collections.defaultdict(list)
    for ev in eventos_fase1_buses:
        bus_raw = ev.get("bus")
        if bus_raw is None or bus_raw == "":
            continue
        try:
            bus_id_int = int(bus_raw)
        except (TypeError, ValueError):
            continue
        eventos_por_bus[bus_id_int].append(ev)
    for bus_id in eventos_por_bus:
        eventos_por_bus[bus_id].sort(key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))

    tarea_a_conductor: Dict[Tuple[Any, int], int] = {}
    for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
        for tid, bus_idx in turno.get("tareas_con_bus", []):
            if tid is not None:
                tarea_a_conductor[(tid, bus_idx)] = c_id
                tarea_a_conductor[(str(tid), bus_idx)] = c_id

    # Mapa viaje_id -> bus_idx (solo ítems de bloque que son viaje comercial, primera ocurrencia)
    viaje_a_bus: Dict[Any, int] = {}
    for id_bus, bloque in enumerate(bloques_bus):
        for viaje in bloque:
            if (viaje.get("evento") or "").strip().lower() in ("vacio", "parada", "recarga"):
                continue
            vid = viaje.get("id") or viaje.get("_tmp_id")
            if vid is not None and vid not in viaje_a_bus and str(vid) not in viaje_a_bus:
                viaje_a_bus[vid] = id_bus
                viaje_a_bus[str(vid)] = id_bus

    # Únicos comerciales = exactamente los del insumo de carga (ni más ni menos)
    comerciales_desde_insumo: List[Dict[str, Any]] = []
    for v in (viajes_comerciales or []):
        vid = v.get("id") or v.get("_tmp_id")
        bus_idx = viaje_a_bus.get(vid) if vid is not None else None
        if bus_idx is None and vid is not None:
            bus_idx = viaje_a_bus.get(str(vid))
        if bus_idx is None:
            # Fallback: buscar por (origen, destino, inicio, fin) en bloques
            for id_bus, bloque in enumerate(bloques_bus):
                for viaje in bloque:
                    if (viaje.get("evento") or "").strip().lower() in ("vacio", "parada", "recarga"):
                        continue
                    if (viaje.get("origen") or "") != (v.get("origen") or ""):
                        continue
                    if (viaje.get("destino") or "") != (v.get("destino") or ""):
                        continue
                    if viaje.get("inicio") != v.get("inicio") or viaje.get("fin") != v.get("fin"):
                        continue
                    bus_idx = id_bus
                    break
                if bus_idx is not None:
                    break
        bus_id = (bus_idx + 1) if bus_idx is not None else ""
        inicio_min = _tiempo_a_minutos(v.get("inicio", 0))
        fin_min = _tiempo_a_minutos(v.get("fin", 0))

        conductor_asignado = None
        if bus_idx is not None and bus_id:
            conductor_asignado = tarea_a_conductor.get((vid, bus_idx)) or (tarea_a_conductor.get((str(vid), bus_idx)) if vid is not None else None)
        if conductor_asignado is None and vid is not None:
            s_vid = str(vid)
            for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                for tid, _ in turno.get("tareas_con_bus", []):
                    if tid == vid or str(tid) == s_vid:
                        conductor_asignado = c_id
                        break
                if conductor_asignado is not None:
                    break
        if conductor_asignado is None and bus_idx is not None:
            for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                for _tid, t_bus_idx in turno.get("tareas_con_bus", []):
                    if t_bus_idx == bus_idx:
                        conductor_asignado = c_id
                        break
                if conductor_asignado is not None:
                    break

        tipo_bus_com = ""
        if bus_idx is not None and bus_idx < len(bloques_bus):
            for viaje_blq in bloques_bus[bus_idx]:
                if (viaje_blq.get("evento") or "").strip().lower() not in ("vacio", "parada", "recarga"):
                    tipo_bus_com = viaje_blq.get("tipo_bus", "") or ""
                    break
        comerciales_desde_insumo.append({
            "evento": "Comercial",
            "bus": bus_id,
            "conductor": conductor_asignado or "",
            "inicio": formatear_hora_deltatime(v.get("inicio", 0)),
            "fin": formatear_hora_deltatime(v.get("fin", 0)),
            "duracion": formatear_hora_deltatime((v.get("fin") or 0) - (v.get("inicio") or 0)),
            "origen": v.get("origen", ""),
            "destino": v.get("destino", ""),
            "linea": v.get("linea", ""),
            "kilometros": v.get("kilometros", 0),
            "desc": v.get("desc", ""),
            "viaje_id": vid,
            "sentido": v.get("sentido", ""),
            "tipo_bus": tipo_bus_com,
        })

    # Regla Fase 3: solo asociar conductor a Vacio/Parada si tiene al menos un Comercial con ese bus.
    conduc_bus_con_comercial: set = set()
    for ev in comerciales_desde_insumo:
        c, b = ev.get("conductor"), ev.get("bus")
        if c and b:
            try:
                conduc_bus_con_comercial.add((int(c), int(b)))
            except (TypeError, ValueError):
                pass

    eventos_fase2_conductores = copy.deepcopy(eventos_fase1_buses)
    for evento in eventos_fase2_conductores:
        t_ev = (str(evento.get("evento", "")) or "").strip().upper()
        if t_ev not in ("COMERCIAL", "VACIO", "PARADA", "RECARGA"):
            continue
        bus_id_evento = evento.get("bus")
        if bus_id_evento in (None, "", 0, "0"):
            continue
        try:
            bus_id_evento = int(bus_id_evento)
        except (TypeError, ValueError):
            continue
        bus_idx_0 = bus_id_evento - 1
        inicio_min = _tiempo_a_minutos(evento.get("inicio", "00:00"))
        fin_min = _tiempo_a_minutos(evento.get("fin", "00:00"))
        conductor_asignado = None
        if fin_min > inicio_min:
            for minuto in range(inicio_min, fin_min):
                conductor_asignado = conductor_por_bus_tiempo.get((bus_id_evento, minuto))
                if conductor_asignado:
                    break
        else:
            conductor_asignado = conductor_por_bus_tiempo.get((bus_id_evento, inicio_min))
        if not conductor_asignado:
            for offset in range(1, 31):
                conductor_asignado = conductor_por_bus_tiempo.get((bus_id_evento, inicio_min - offset))
                if conductor_asignado:
                    break
        if not conductor_asignado:
            for offset in range(0, 31):
                conductor_asignado = conductor_por_bus_tiempo.get((bus_id_evento, fin_min + offset))
                if conductor_asignado:
                    break
        if conductor_asignado:
            # Solo asociar a Vacio/Parada si el conductor tiene Comercial con ese bus (regla Fase 3).
            if t_ev in ("VACIO", "PARADA"):
                if (conductor_asignado, bus_id_evento) not in conduc_bus_con_comercial:
                    conductor_asignado = None
            if conductor_asignado:
                evento["conductor"] = conductor_asignado
        elif t_ev in ("VACIO", "PARADA"):
            # Vacíos y paradas: solo asignar si el conductor tiene Comercial con ese bus (regla Fase 3).
            for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                for _tid, t_bus_idx in turno.get("tareas_con_bus", []):
                    if t_bus_idx == bus_idx_0 and (c_id, bus_id_evento) in conduc_bus_con_comercial:
                        conductor_asignado = c_id
                        break
                if conductor_asignado is not None:
                    break
            if conductor_asignado is not None:
                evento["conductor"] = conductor_asignado
        elif evento.get("evento") == "Comercial":
            vid = evento.get("viaje_id")
            if vid is not None:
                # 1) Buscar conductor por (viaje_id, bus_idx) como en Fase 2/3
                conductor_asignado = tarea_a_conductor.get((vid, bus_idx_0)) or tarea_a_conductor.get((str(vid), bus_idx_0))
            # 2) Si falla, usar matching suave contra viajes_comerciales (por origen/destino/horario)
            if conductor_asignado is None and evento.get("origen") is not None and evento.get("destino") is not None:
                for v in (viajes_comerciales or []):
                    if (v.get("origen") or "").strip() != (evento.get("origen") or "").strip():
                        continue
                    if (v.get("destino") or "").strip() != (evento.get("destino") or "").strip():
                        continue
                    vi = _tiempo_a_minutos(v.get("inicio", 0))
                    vf = _tiempo_a_minutos(v.get("fin", 0))
                    if abs(vi - inicio_min) <= 15 and abs(vf - fin_min) <= 15:
                        vid = v.get("id") or v.get("_tmp_id")
                        if vid is not None:
                            conductor_asignado = tarea_a_conductor.get((vid, bus_idx_0)) or tarea_a_conductor.get((str(vid), bus_idx_0))
                        break
            # 3) Si aún no hay conductor, buscar al conductor de ese viaje en cualquier bus (ignorar bus_idx)
            if conductor_asignado is None and vid is not None:
                s_vid = str(vid)
                for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                    for tid, _t_bus_idx in turno.get("tareas_con_bus", []):
                        if tid == vid or str(tid) == s_vid:
                            conductor_asignado = c_id
                            break
                    if conductor_asignado is not None:
                        break
            # 4) Fallback por (bus + rango horario) en todos los turnos
            if conductor_asignado is None:
                for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                    for tid, t_bus_idx in turno.get("tareas_con_bus", []):
                        if t_bus_idx != bus_idx_0:
                            continue
                        v = mapa_viajes.get(tid) or mapa_viajes.get(str(tid))
                        if not v:
                            meta = metadata_tareas.get(tid) or metadata_tareas.get(str(tid))
                            v = meta.get("viaje") if isinstance(meta, dict) else None
                        if v and isinstance(v, dict):
                            vi = _tiempo_a_minutos(v.get("inicio", 0))
                            vf = _tiempo_a_minutos(v.get("fin", 0))
                            if abs(vi - inicio_min) <= 15 and abs(vf - fin_min) <= 15:
                                conductor_asignado = c_id
                                break
                    if conductor_asignado is not None:
                        break
            # 5) Fallback último: cualquier conductor que tenga tarea en este bus (garantizar 100% comerciales)
            if conductor_asignado is None:
                for c_id, turno in enumerate(turnos_seleccionados or [], start=1):
                    for _tid, t_bus_idx in turno.get("tareas_con_bus", []):
                        if t_bus_idx == bus_idx_0:
                            conductor_asignado = c_id
                            break
                    if conductor_asignado is not None:
                        break
            if conductor_asignado is not None:
                evento["conductor"] = conductor_asignado

    eventos_fase2_conductores.extend(comerciales_desde_insumo)

    # Asignar Vacíos depot→nodo sin conductor a conductores cuyo primer Comercial sale de ese nodo
    # (prioridad Vacío sobre Desplazamiento).
    comerciales_por_cond: Dict[int, List[Dict]] = {}
    for ev in comerciales_desde_insumo:
        cid = ev.get("conductor")
        if not cid:
            continue
        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            continue
        comerciales_por_cond.setdefault(cid_int, []).append(ev)
    for cid_int, evs in comerciales_por_cond.items():
        primer_com = min(evs, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)) or 99999)
        bus_id = primer_com.get("bus")
        if not bus_id:
            continue
        try:
            bus_int = int(bus_id)
        except (TypeError, ValueError):
            continue
        if (cid_int, bus_int) not in conduc_bus_con_comercial:
            continue
        origen_com = (primer_com.get("origen") or "").strip()
        if not origen_com or _origen_es_deposito_estricto(origen_com, gestor):
            continue
        ini_com = _tiempo_a_minutos(primer_com.get("inicio", 0))
        if ini_com is None:
            continue
        for vac in eventos_fase2_conductores:
            if (vac.get("evento") or "").strip() != "Vacio" or vac.get("conductor"):
                continue
            bv = vac.get("bus")
            if bv in (None, "", 0, "0"):
                continue
            try:
                if int(bv) != bus_int:
                    continue
            except (TypeError, ValueError):
                continue
            o_v = (vac.get("origen") or "").strip()
            d_v = (vac.get("destino") or "").strip()
            if not _es_deposito(o_v, deposito) or (d_v or "").upper() != (origen_com or "").upper():
                continue
            fin_v = _tiempo_a_minutos(vac.get("fin", 0))
            if fin_v is not None and fin_v <= ini_com + 2:
                vac["conductor"] = cid_int
                break

    for id_conductor, turno in enumerate(turnos_seleccionados or [], start=1):
        tareas_del_turno = turno.get("tareas_con_bus", [])
        if not tareas_del_turno:
            continue
        primera_tarea_id, primer_bus_idx = tareas_del_turno[0]
        primer_viaje_turno = mapa_viajes.get(primera_tarea_id) or mapa_viajes.get(str(primera_tarea_id))
        ultima_tarea_id, ultimo_bus_idx = tareas_del_turno[-1]
        ultimo_viaje_turno = mapa_viajes.get(ultima_tarea_id) or mapa_viajes.get(str(ultima_tarea_id))

        inicio_turno = turno.get("inicio", 0)
        fin_turno = turno.get("fin", 0)

        # Ajustar InS al primer evento real del conductor (vacío o comercial),
        # para que la toma de 15 min termine justo cuando comienza ese evento.
        eventos_cond_existentes = [
            e for e in eventos_fase2_conductores
            if e.get("conductor") == id_conductor
        ]
        if eventos_cond_existentes:
            primer_inicio_ev = min(
                _tiempo_a_minutos(e.get("inicio", 0)) for e in eventos_cond_existentes
            )
        else:
            primer_inicio_ev = inicio_turno
        ins_fin_min = max(0, primer_inicio_ev)
        ins_inicio_min = max(0, ins_fin_min - tiempo_toma)
        ins_evento_actual = {
            "evento": "InS",
            "bus": "",
            "conductor": id_conductor,
            "inicio": formatear_hora_deltatime(ins_inicio_min),
            "fin": formatear_hora_deltatime(ins_fin_min),
            "duracion": formatear_hora_deltatime(ins_fin_min - ins_inicio_min),
            "origen": deposito,
            "destino": deposito,
            "linea": "",
            "kilometros": 0,
            "desc": "Inicio de Jornada (Toma)",
        }
        eventos_fase2_conductores.append(ins_evento_actual)

        punto_fin = (turno.get("punto_fin_turno") or deposito).strip()
        # Nodo real donde el conductor empieza a trabajar: primer evento ya asignado
        nodo_inicio_real = ""
        if eventos_cond_existentes:
            primer_ev_real = min(
                eventos_cond_existentes, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0))
            )
            nodo_inicio_real = (primer_ev_real.get("origen") or primer_ev_real.get("destino") or "").strip()

        # Determinar origen_primero para evaluar conexión depósito→origen:
        # 1) Preferir nodo_inicio_real si no es depósito (ej. LOS TILOS cuando el primer evento es Parada allí).
        # 2) Si no, usar el primer Comercial por horario.
        # 3) Como último recurso, usar el origen del primer viaje del turno.
        origen_primero = ""
        if nodo_inicio_real and not _origen_es_deposito_estricto(nodo_inicio_real, gestor):
            origen_primero = nodo_inicio_real
        elif comerciales_desde_insumo:
            comerciales_cond = [ev for ev in comerciales_desde_insumo if ev.get("conductor") == id_conductor]
            if comerciales_cond:
                primer_comercial_cond = min(comerciales_cond, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))
                origen_primero = (primer_comercial_cond.get("origen") or "").strip()
        if not origen_primero and primer_viaje_turno:
            origen_primero = (primer_viaje_turno.get("origen", "") or "").strip()
        if origen_primero and not _origen_es_deposito_estricto(origen_primero, gestor):
            # Prioridad: Vacío > Desplazamiento. Si ya existe un Vacío o Desplazamiento
            # depot→origen_primero para este conductor, no crear Desplazamiento.
            hay_despl_conexion = False
            primer_comercial_min = None
            if comerciales_desde_insumo:
                comerciales_cond = [
                    ev for ev in comerciales_desde_insumo if ev.get("conductor") == id_conductor
                ]
                if comerciales_cond:
                    primer_comercial_min = min(
                        _tiempo_a_minutos(ev.get("inicio", 0)) for ev in comerciales_cond
                    )
            for ev in eventos_cond_existentes:
                if (ev.get("evento") or "").strip() not in ("Vacio", "Desplazamiento"):
                    continue
                o = (ev.get("origen") or "").strip()
                d = (ev.get("destino") or "").strip()
                if not _es_deposito(o, deposito) or d.strip() != origen_primero:
                    continue
                ini_v = _tiempo_a_minutos(ev.get("inicio", 0))
                fin_v = _tiempo_a_minutos(ev.get("fin", 0))
                if ini_v >= ins_fin_min - 1 and (primer_comercial_min is None or fin_v <= primer_comercial_min + 1):
                    hay_despl_conexion = True
                    break
                        
            if not hay_despl_conexion:
                # Solo crear Desplazamiento cuando NO existe Vacío depot→origen.
                # Si la conexión está habilitada, ajustar InS y agregar Desplazamiento.
                if comerciales_desde_insumo and comerciales_cond:
                    objetivo_inicio = primer_comercial_min
                else:
                    objetivo_inicio = primer_inicio_ev
                if objetivo_inicio is None:
                    objetivo_inicio = ins_fin_min

                hab, t_despl = gestor.buscar_info_desplazamiento(
                    deposito, origen_primero, objetivo_inicio
                )
                if hab and t_despl and t_despl > 0 and objetivo_inicio >= t_despl:
                    # Nuevo fin de InS = inicio_comercial - t_despl
                    nuevo_ins_fin = objetivo_inicio - t_despl
                    nuevo_ins_ini = max(0, nuevo_ins_fin - tiempo_toma)
                    if nuevo_ins_fin > nuevo_ins_ini:
                        ins_evento_actual["inicio"] = formatear_hora_deltatime(nuevo_ins_ini)
                        ins_evento_actual["fin"] = formatear_hora_deltatime(nuevo_ins_fin)
                        ins_evento_actual["duracion"] = formatear_hora_deltatime(
                            nuevo_ins_fin - nuevo_ins_ini
                        )
                        inicio_despl = nuevo_ins_fin
                        fin_despl = objetivo_inicio
                        eventos_fase2_conductores.append({
                            "evento": "Desplazamiento",
                            "bus": "",
                            "conductor": id_conductor,
                            "inicio": formatear_hora_deltatime(inicio_despl),
                            "fin": formatear_hora_deltatime(fin_despl),
                            "duracion": formatear_hora_deltatime(fin_despl - inicio_despl),
                            "origen": deposito,
                            "destino": origen_primero,
                            "linea": "",
                            "kilometros": 0,
                            "desc": f"Desplazamiento {deposito}->{origen_primero}",
                        })

        if ultimo_viaje_turno and not _es_deposito(punto_fin, deposito):
            # Si ya existe un Vacio/Desplazamiento nodo→depósito para este
            # conductor (ej. PIE ANDINO -> Deposito Pie Andino), no crear un
            # Desplazamiento adicional: se usa ese Vacio para volver y luego FnS.
            hay_retorno = False
            for ev in eventos_cond_existentes:
                tipo_ev = (ev.get("evento") or "").strip()
                if tipo_ev not in ("Vacio", "Desplazamiento"):
                    continue
                o = (ev.get("origen") or "").strip()
                d = (ev.get("destino") or "").strip()
                if o == punto_fin and _es_deposito(d, deposito):
                    hay_retorno = True
                    break
                            
            if not hay_retorno:
                valido_relay, _, t_node_to_dep = _es_relevo_valido(punto_fin, deposito, gestor)
                if valido_relay:
                    hab, t_despl = gestor.buscar_info_desplazamiento(
                        punto_fin, deposito, ultimo_viaje_turno["fin"]
                    )
                    if hab and t_despl:
                        # Encadenar: el desplazamiento de vuelta al depósito debe
                        # comenzar exactamente cuando termina el último evento real
                        # del conductor (comercial/vacío/parada), sin huecos.
                        if eventos_cond_existentes:
                            fin_ultimo_ev = max(
                                _tiempo_a_minutos(e.get("fin", 0)) for e in eventos_cond_existentes
                            )
                        else:
                            fin_ultimo_ev = ultimo_viaje_turno["fin"]
                        inicio_despl = fin_ultimo_ev
                        fin_despl = inicio_despl + t_despl
                        eventos_fase2_conductores.append({
                            "evento": "Desplazamiento",
                            "bus": "",
                            "conductor": id_conductor,
                            "inicio": formatear_hora_deltatime(inicio_despl),
                            "fin": formatear_hora_deltatime(fin_despl),
                            "duracion": formatear_hora_deltatime(t_despl),
                            "origen": punto_fin,
                            "destino": deposito,
                            "linea": "",
                            "kilometros": 0,
                            "desc": f"Desplazamiento {punto_fin}->{deposito}",
                        })

        eventos_fase2_conductores.append({
            "evento": "FnS",
            "bus": "",
            "conductor": id_conductor,
            "inicio": formatear_hora_deltatime(fin_turno),
            "fin": formatear_hora_deltatime(fin_turno),
            "duracion": "00:00",
            "origen": deposito,
            "destino": deposito,
            "linea": "",
            "kilometros": 0,
            "desc": "Fin de Jornada (Deja)",
        })

    # Encadenar eventos por conductor: sin huecos temporales dentro de la jornada.
    # No se modifican horarios de Comerciales ni Vacíos de Fase 1; si hay huecos
    # se rellenan con Paradas sintéticas (descanso) usando el nodo destino anterior.
    eventos_por_conductor: Dict[Any, List[Dict[str, Any]]] = collections.defaultdict(list)
    for ev in eventos_fase2_conductores:
        cid = ev.get("conductor")
        if cid:
            eventos_por_conductor[cid].append(ev)

    nuevas_paradas: List[Dict[str, Any]] = []
    nuevos_desplazamientos_finales: List[Dict[str, Any]] = []
    for cid, lista in eventos_por_conductor.items():
        if not lista:
            continue
        lista_ordenada = sorted(lista, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))

        # Asegurar Desplazamiento explícito nodo→depósito antes de FnS cuando
        # el último nodo no es depósito (ej. PIE ANDINO → Deposito Pie Andino).
        for i in range(1, len(lista_ordenada)):
            prev = lista_ordenada[i - 1]
            curr = lista_ordenada[i]
            if (curr.get("evento") or "").strip() != "FnS":
                continue
            if (prev.get("evento") or "").strip() == "Desplazamiento":
                continue
            dest_prev = (prev.get("destino") or prev.get("origen") or "").strip()
            if not dest_prev or _es_deposito(dest_prev, deposito):
                    continue
            if not _es_deposito(curr.get("origen", ""), deposito) or not _es_deposito(
                curr.get("destino", ""), deposito
            ):
                    continue
            fin_prev = _tiempo_a_minutos(prev.get("fin", 0))
            if fin_prev is None:
                    continue
            hab_d, t_d = gestor.buscar_info_desplazamiento(dest_prev, deposito, fin_prev)
            if not (hab_d and t_d):
                continue
            inicio_despl = fin_prev
            fin_despl = inicio_despl + t_d
            nuevos_desplazamientos_finales.append({
                "evento": "Desplazamiento",
                                    "bus": "",
                                    "conductor": cid,
                "inicio": formatear_hora_deltatime(inicio_despl),
                "fin": formatear_hora_deltatime(fin_despl),
                "duracion": formatear_hora_deltatime(t_d),
                "origen": dest_prev,
                "destino": deposito,
                "linea": "",
                "kilometros": 0,
                "desc": f"Desplazamiento {dest_prev}->{deposito}",
            })
            # Mover FnS al final del desplazamiento
            curr["inicio"] = formatear_hora_deltatime(fin_despl)
            curr["fin"] = formatear_hora_deltatime(fin_despl)
            curr["duracion"] = "00:00"

        # Encadenado temporal robusto: usar el último fin efectivo, no solo pares consecutivos,
        # para evitar crear Paradas ficticias cuando hay eventos solapados (ej. Desplazamiento
        # más corto anidado dentro de un Comercial más largo).
        ultimo_ev = min(lista_ordenada, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))
        ultimo_fin = _tiempo_a_minutos(ultimo_ev.get("fin", 0))
        for curr in sorted(lista_ordenada[1:], key=lambda e: _tiempo_a_minutos(e.get("inicio", 0))):
            ini_curr = _tiempo_a_minutos(curr.get("inicio", 0))
            fin_curr = _tiempo_a_minutos(curr.get("fin", 0))
            if ini_curr is None or fin_curr is None or ultimo_fin is None:
                # Actualizar último fin de todas formas si es mayor
                if fin_curr is not None and (ultimo_fin is None or fin_curr > ultimo_fin):
                    ultimo_fin = fin_curr
                    ultimo_ev = curr
                continue
            if ini_curr > ultimo_fin:
                gap = ini_curr - ultimo_fin
                nuevas_paradas.append({
                    "evento": "Parada",
                            "bus": "",
                            "conductor": cid,
                    "inicio": formatear_hora_deltatime(ultimo_fin),
                    "fin": formatear_hora_deltatime(ini_curr),
                    "duracion": formatear_hora_deltatime(gap),
                    "origen": ultimo_ev.get("destino", "") or ultimo_ev.get("origen", ""),
                    "destino": ultimo_ev.get("destino", "") or ultimo_ev.get("origen", ""),
                    "linea": "",
                    "kilometros": 0,
                    "desc": "Parada/descanso",
                })
            # Avanzar último evento si este termina más tarde
            if fin_curr >= ultimo_fin:
                ultimo_fin = fin_curr
                ultimo_ev = curr

    if nuevas_paradas:
        eventos_fase2_conductores.extend(nuevas_paradas)
    if nuevos_desplazamientos_finales:
        eventos_fase2_conductores.extend(nuevos_desplazamientos_finales)

    # Limpieza final por conductor:
    # 1) Eliminar Paradas "flotantes" sin bus que queden contenidas o solapadas
    #    dentro de otro evento del mismo conductor (evitar dobles capas).
    # 2) Eliminar Vacíos de depósito→nodo asignados al conductor cuando el
    #    evento anterior ya lo deja en ese nodo (el vacío es solo reposicionamiento
    #    de bus, no del conductor).
    eventos_por_conductor_limpio: Dict[Any, List[int]] = collections.defaultdict(list)
    for idx, ev in enumerate(eventos_fase2_conductores):
        cid = ev.get("conductor")
        if cid:
            eventos_por_conductor_limpio[cid].append(idx)
    indices_a_eliminar: Set[int] = set()
    for cid, idxs in eventos_por_conductor_limpio.items():
        # Ordenar índices de este conductor por inicio
        idxs_ordenados = sorted(
            idxs, key=lambda k: _tiempo_a_minutos(eventos_fase2_conductores[k].get("inicio", 0))
        )

        # 1) Paradas flotantes sin bus que se solapan con otros eventos
        for i in idxs_ordenados:
            ev = eventos_fase2_conductores[i]
            if (ev.get("evento") or "").strip() != "Parada":
                continue
            bus_ev = ev.get("bus")
            if bus_ev not in (None, "", 0, "0"):
                continue
            ini_p = _tiempo_a_minutos(ev.get("inicio", 0))
            fin_p = _tiempo_a_minutos(ev.get("fin", 0))
            if ini_p is None or fin_p is None:
                continue
            for j in idxs_ordenados:
                if j == i:
                    continue
                otro = eventos_fase2_conductores[j]
                if (otro.get("evento") or "").strip() == "Parada" and (
                    otro.get("bus") in (None, "", 0, "0")
                ):
                    continue
                ini_o = _tiempo_a_minutos(otro.get("inicio", 0))
                fin_o = _tiempo_a_minutos(otro.get("fin", 0))
                if ini_o is None or fin_o is None:
                    continue
                if max(ini_o, ini_p) < min(fin_o, fin_p):
                    indices_a_eliminar.add(i)
                    break

        # 2) Vacíos depósito→nodo innecesarios para el conductor:
        #    patrón: ... [dest=N] seguido por Vacio(deposito->N) pequeño.
        for a, b in zip(idxs_ordenados, idxs_ordenados[1:]):
            prev = eventos_fase2_conductores[a]
            curr = eventos_fase2_conductores[b]
            if (curr.get("evento") or "").strip() != "Vacio":
                continue
            bus_curr = curr.get("bus")
            if bus_curr in (None, "", 0, "0"):
                continue
            origen_curr = (curr.get("origen") or "").strip()
            destino_curr = (curr.get("destino") or "").strip()
            dest_prev = (prev.get("destino") or prev.get("origen") or "").strip()
            if not dest_prev or not destino_curr:
                continue
            if not _es_deposito(origen_curr, deposito):
                continue
            if _es_deposito(dest_prev, deposito):
                continue
            # Si el destino del vacío coincide con el nodo donde ya está el
            # conductor, este vacío es solo reposicionamiento del bus. Quitar
            # el evento de la vista de conductor.
            if destino_curr == dest_prev:
                indices_a_eliminar.add(b)

        # 3) Desplazamientos depósito→nodo redundantes cuando ya existe un Vacío
        #    depósito→mismo nodo en el mismo intervalo. Se prioriza Vacío sobre
        #    Desplazamiento: el conductor usa el Vacío del bus (si tiene Comercial
        #    con ese bus); Desplazamiento solo cuando no exista Vacío.
        for i in idxs_ordenados:
            ev = eventos_fase2_conductores[i]
            if (ev.get("evento") or "").strip() != "Desplazamiento":
                continue
            origen_d = (ev.get("origen") or "").strip()
            destino_d = (ev.get("destino") or "").strip()
            if not _es_deposito(origen_d, deposito):
                continue
            if not destino_d or _es_deposito(destino_d, deposito):
                continue
            cid_d = ev.get("conductor")
            ini_d = _tiempo_a_minutos(ev.get("inicio", 0))
            fin_d = _tiempo_a_minutos(ev.get("fin", 0))
            if ini_d is None or fin_d is None:
                continue
            for j in idxs_ordenados:
                if j == i:
                    continue
                otro = eventos_fase2_conductores[j]
                if (otro.get("evento") or "").strip() != "Vacio":
                    continue
                if otro.get("conductor") != cid_d:
                    continue
                origen_v = (otro.get("origen") or "").strip()
                destino_v = (otro.get("destino") or "").strip()
                if not _es_deposito(origen_v, deposito) or destino_v != destino_d:
                    continue
                ini_v = _tiempo_a_minutos(otro.get("inicio", 0))
                fin_v = _tiempo_a_minutos(otro.get("fin", 0))
                if ini_v is None or fin_v is None:
                    continue
                # Si se solapan en el tiempo, el Desplazamiento es redundante:
                # el conductor ya usa el Vacío del bus.
                if max(ini_v, ini_d) < min(fin_v, fin_d):
                    indices_a_eliminar.add(i)
                    break

        # 4) Desplazamientos duplicados exactos (mismo intervalo y mismo origen/destino)
        vistos_disp = set()
        for i in idxs_ordenados:
            ev = eventos_fase2_conductores[i]
            if (ev.get("evento") or "").strip() != "Desplazamiento":
                continue
            clave = (
                (ev.get("origen") or "").strip(),
                (ev.get("destino") or "").strip(),
                (ev.get("inicio") or "").strip(),
                (ev.get("fin") or "").strip(),
            )
            if clave in vistos_disp:
                indices_a_eliminar.add(i)
            else:
                vistos_disp.add(clave)
    if indices_a_eliminar:
        eventos_fase2_conductores = [
            ev for idx, ev in enumerate(eventos_fase2_conductores) if idx not in indices_a_eliminar
        ]

    # Ajuste de encadenamiento de nodos tras Parada en depósito antes de un Comercial
    # de relevo: si el conductor está en el depósito y el siguiente Comercial
    # comienza en un nodo distinto con desplazamiento configurado, usar la cola
    # de la Parada para insertar ese Desplazamiento y evitar "teletransportes".
    eventos_por_cid_enc: Dict[Any, List[Dict[str, Any]]] = collections.defaultdict(list)
    for ev in eventos_fase2_conductores:
        cid = ev.get("conductor")
        if cid:
            eventos_por_cid_enc[cid].append(ev)
    nuevos_despl_relevo: List[Dict[str, Any]] = []
    for cid, lista in eventos_por_cid_enc.items():
        ordenada = sorted(lista, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))
        for i in range(len(ordenada) - 1):
            ev = ordenada[i]
            nxt = ordenada[i + 1]
            if (ev.get("evento") or "").strip() != "Parada":
                continue
            if (nxt.get("evento") or "").strip() != "Comercial":
                continue
            nodo_parada = (ev.get("destino") or ev.get("origen") or "").strip()
            origen_next = (nxt.get("origen") or "").strip()
            if not nodo_parada or not origen_next:
                continue
            if not _es_deposito(nodo_parada, deposito):
                continue
            if _origen_es_deposito_estricto(origen_next, gestor):
                    continue
            ini_p = _tiempo_a_minutos(ev.get("inicio", 0))
            fin_p = _tiempo_a_minutos(ev.get("fin", 0))
            ini_next = _tiempo_a_minutos(nxt.get("inicio", 0))
            if ini_p is None or fin_p is None or ini_next is None:
                continue
            if ini_next <= fin_p:
                gap_total = ini_next - ini_p
            else:
                gap_total = ini_next - fin_p
            if gap_total <= 0:
                        continue
            hab, t_despl = gestor.buscar_info_desplazamiento(deposito, origen_next, fin_p)
            if not (hab and t_despl and t_despl > 0):
                    continue
            if t_despl > (ini_next - ini_p):
                    continue
            nuevo_fin_parada = ini_next - t_despl
            if nuevo_fin_parada <= ini_p:
                continue
            ev["fin"] = formatear_hora_deltatime(nuevo_fin_parada)
            ev["duracion"] = formatear_hora_deltatime(nuevo_fin_parada - ini_p)
            nuevos_despl_relevo.append({
                "evento": "Desplazamiento",
                "bus": "",
                "conductor": cid,
                "inicio": formatear_hora_deltatime(nuevo_fin_parada),
                "fin": formatear_hora_deltatime(ini_next),
                "duracion": formatear_hora_deltatime(ini_next - nuevo_fin_parada),
                "origen": deposito,
                "destino": origen_next,
                "linea": "",
                "kilometros": 0,
                "desc": f"Desplazamiento {deposito}->{origen_next}",
            })
    if nuevos_despl_relevo:
        eventos_fase2_conductores.extend(nuevos_despl_relevo)

    # 3) Asegurar conexión depósito→primer nodo: si el primer evento real del conductor
    #    es un Comercial con origen no depósito y no hay Vacio/Desplazamiento entre InS
    #    y ese Comercial, agregar Desplazamiento explícito (evita "falta conexión" dep→PIE ANDINO, etc.).
    eventos_por_cid: Dict[Any, List[Dict[str, Any]]] = collections.defaultdict(list)
    for ev in eventos_fase2_conductores:
        cid = ev.get("conductor")
        if cid:
            eventos_por_cid[cid].append(ev)
    desplazamientos_iniciales: List[Dict[str, Any]] = []
    for cid, lista in eventos_por_cid.items():
        ordenada = sorted(lista, key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))
        ins_ev = None
        primer_real = None
        for ev in ordenada:
            tipo = (ev.get("evento") or "").strip()
            if tipo == "InS":
                ins_ev = ev
                continue
            if tipo in ("FnS",):
                continue
            primer_real = ev
            break
        if not ins_ev or not primer_real:
            continue
        # Tiempo actual de fin de InS (se puede reajustar si añadimos Desplazamiento)
        ins_fin = _tiempo_a_minutos(ins_ev.get("fin", 0))
        origen_primer = (primer_real.get("origen") or "").strip()
        if not origen_primer or _origen_es_deposito_estricto(origen_primer, gestor):
            continue
        # ¿Hay ya un Vacio/Desplazamiento dep→origen_primer entre InS y primer_real?
        ini_primer = _tiempo_a_minutos(primer_real.get("inicio", 0))
        hay_conexion = False
        for ev in ordenada:
            t = (ev.get("evento") or "").strip()
            if t not in ("Vacio", "Desplazamiento"):
                continue
            o = (ev.get("origen") or "").strip()
            d = (ev.get("destino") or "").strip()
            if _es_deposito(o, deposito) and (d == origen_primer or (d or "").upper() == (origen_primer or "").upper()):
                ev_ini = _tiempo_a_minutos(ev.get("inicio", 0))
                ev_fin = _tiempo_a_minutos(ev.get("fin", 0))
                if ev_ini is not None and ev_fin is not None and ins_fin is not None and ini_primer is not None:
                    if ev_ini >= ins_fin - 2 and ev_fin <= ini_primer + 2:
                        hay_conexion = True
                        break
        if hay_conexion:
            continue

        # Antes de crear Desplazamiento: buscar Vacío depot→origen_primer SIN conductor
        # que tenga bus con Comercial para este conductor. Asignarlo y no crear Desplazamiento.
        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            cid_int = None
        if cid_int is not None:
            for ev in eventos_fase2_conductores:
                if (ev.get("evento") or "").strip() != "Vacio":
                    continue
                if ev.get("conductor"):
                    continue
                bus_v = ev.get("bus")
                if bus_v in (None, "", 0, "0"):
                    continue
                try:
                    bus_v_int = int(bus_v)
                except (TypeError, ValueError):
                    continue
                if (cid_int, bus_v_int) not in conduc_bus_con_comercial:
                    continue
                o_v = (ev.get("origen") or "").strip()
                d_v = (ev.get("destino") or "").strip()
                if not _es_deposito(o_v, deposito):
                    continue
                if (d_v or "").upper() != (origen_primer or "").upper():
                    continue
                ev_ini = _tiempo_a_minutos(ev.get("inicio", 0))
                ev_fin = _tiempo_a_minutos(ev.get("fin", 0))
                if ev_ini is not None and ev_fin is not None and ins_fin is not None and ini_primer is not None:
                    # Ventana amplia: Vacío debe conectar InS con primer evento real
                    if ev_ini <= ini_primer + 5 and ev_fin >= ins_fin - 5:
                        ev["conductor"] = cid
                        hay_conexion = True
                        break
        if hay_conexion:
            continue

        # Buscar Vacío depot→origen_primer YA asignado a este conductor.
        for ev in eventos_fase2_conductores:
            if (ev.get("evento") or "").strip() != "Vacio":
                continue
            if ev.get("conductor") != cid:
                continue
            o_v = (ev.get("origen") or "").strip()
            d_v = (ev.get("destino") or "").strip()
            if not _es_deposito(o_v, deposito):
                continue
            if (d_v or "").upper() != (origen_primer or "").upper():
                continue
            ev_ini = _tiempo_a_minutos(ev.get("inicio", 0))
            ev_fin = _tiempo_a_minutos(ev.get("fin", 0))
            if ev_ini is not None and ev_fin is not None and ins_fin is not None and ini_primer is not None:
                if ev_ini >= ins_fin - 2 and ev_fin <= ini_primer + 2:
                    hay_conexion = True
                    break
        if hay_conexion:
            continue

        # Crear un Desplazamiento dep→origen_primer ANTES del primer evento, ajustando
        # también el InS para que la toma + desplazamiento terminen justo al inicio
        # del primer evento real (sin solaparse).
        if ini_primer is None:
            continue
        hab, t_despl = gestor.buscar_info_desplazamiento(deposito, origen_primer, ini_primer)
        if not (hab and t_despl and t_despl > 0):
            continue
        # Necesitamos que haya al menos t_despl minutos antes del primer evento
        if ini_primer < t_despl:
            continue
        # Fin deseado de InS = inicio_primer - t_despl
        tiempo_toma = getattr(gestor, "tiempo_toma", 0) or 0
        nuevo_ins_fin = ini_primer - t_despl
        nuevo_ins_ini = max(0, nuevo_ins_fin - tiempo_toma)
        if nuevo_ins_fin <= nuevo_ins_ini:
            continue

        # Reajustar evento InS existente
        ins_ev["inicio"] = formatear_hora_deltatime(nuevo_ins_ini)
        ins_ev["fin"] = formatear_hora_deltatime(nuevo_ins_fin)
        ins_ev["duracion"] = formatear_hora_deltatime(nuevo_ins_fin - nuevo_ins_ini)

        # Desplazamiento exacto entre fin de InS y primer evento
        inicio_despl = nuevo_ins_fin
        fin_despl = ini_primer
        desplazamientos_iniciales.append({
            "evento": "Desplazamiento",
            "bus": "",
            "conductor": cid,
            "inicio": formatear_hora_deltatime(inicio_despl),
            "fin": formatear_hora_deltatime(fin_despl),
            "duracion": formatear_hora_deltatime(fin_despl - inicio_despl),
            "origen": deposito,
            "destino": origen_primer,
            "linea": "",
            "kilometros": 0,
            "desc": f"Desplazamiento {deposito}->{origen_primer}",
        })
    if desplazamientos_iniciales:
        eventos_fase2_conductores.extend(desplazamientos_iniciales)

    # Deduplicación global final de Desplazamientos exactos por conductor
    vistos_glob_disp = set()
    eventos_filtrados: List[Dict[str, Any]] = []
    for ev in eventos_fase2_conductores:
        if (ev.get("evento") or "").strip() != "Desplazamiento":
            eventos_filtrados.append(ev)
            continue
        clave = (
            ev.get("conductor"),
            (ev.get("origen") or "").strip(),
            (ev.get("destino") or "").strip(),
            (ev.get("inicio") or "").strip(),
            (ev.get("fin") or "").strip(),
        )
        if clave in vistos_glob_disp:
                continue
        vistos_glob_disp.add(clave)
        eventos_filtrados.append(ev)
    eventos_fase2_conductores = eventos_filtrados

    # Eliminar Paradas "colgando" inmediatamente después de InS o
    # inmediatamente antes de FnS: no debe haber Parada entre InS y el
    # primer evento real, ni entre el último evento real y FnS.
    eventos_por_cid_clean2: Dict[Any, List[int]] = collections.defaultdict(list)
    for idx, ev in enumerate(eventos_fase2_conductores):
        cid = ev.get("conductor")
        if cid:
            eventos_por_cid_clean2[cid].append(idx)
    indices_a_eliminar_paradas_extremos: Set[int] = set()
    for cid, idxs in eventos_por_cid_clean2.items():
        idxs_ordenados = sorted(
            idxs, key=lambda k: _tiempo_a_minutos(eventos_fase2_conductores[k].get("inicio", 0))
        )
        # identificar primer y último evento REAL (no InS/FnS)
        primer_real_idx = None
        ultimo_real_idx = None
        for i in idxs_ordenados:
            t = (eventos_fase2_conductores[i].get("evento") or "").strip()
            if t in ("InS", "FnS", "Parada"):
                continue
            primer_real_idx = i
            break
        for i in reversed(idxs_ordenados):
            t = (eventos_fase2_conductores[i].get("evento") or "").strip()
            if t in ("InS", "FnS", "Parada"):
                continue
            ultimo_real_idx = i
            break
        if primer_real_idx is None or ultimo_real_idx is None:
            continue
        ini_primer_real = _tiempo_a_minutos(eventos_fase2_conductores[primer_real_idx].get("inicio", 0))
        fin_ultimo_real = _tiempo_a_minutos(eventos_fase2_conductores[ultimo_real_idx].get("fin", 0))
        if ini_primer_real is None or fin_ultimo_real is None:
            continue
        for i in idxs_ordenados:
            ev = eventos_fase2_conductores[i]
            if (ev.get("evento") or "").strip() != "Parada":
                continue
            ini_p = _tiempo_a_minutos(ev.get("inicio", 0))
            fin_p = _tiempo_a_minutos(ev.get("fin", 0))
            if ini_p is None or fin_p is None:
                continue
            # Parada completamente antes del primer real -> eliminar
            if fin_p <= ini_primer_real:
                indices_a_eliminar_paradas_extremos.add(i)
                continue
            # Parada completamente después del último real -> eliminar
            if ini_p >= fin_ultimo_real:
                indices_a_eliminar_paradas_extremos.add(i)
                continue
    if indices_a_eliminar_paradas_extremos:
        eventos_fase2_conductores = [
            ev for idx, ev in enumerate(eventos_fase2_conductores) if idx not in indices_a_eliminar_paradas_extremos
        ]

    # Eliminar Paradas con bus antes del primer Comercial del conductor.
    # Ejemplo: bus esperando en LOS TILOS antes de que llegue el conductor
    # desde el depósito; esa espera no debe contarse como Parada del conductor.
    eventos_por_cid_pre: Dict[Any, List[int]] = collections.defaultdict(list)
    for idx, ev in enumerate(eventos_fase2_conductores):
        cid = ev.get("conductor")
        if cid:
            eventos_por_cid_pre[cid].append(idx)
    indices_a_eliminar_paradas_pre: Set[int] = set()
    for cid, idxs in eventos_por_cid_pre.items():
        comerciales = [
            eventos_fase2_conductores[i]
            for i in idxs
            if (eventos_fase2_conductores[i].get("evento") or "").strip() == "Comercial"
        ]
        if not comerciales:
            continue
        t_primer_com = min(
            _tiempo_a_minutos(ev.get("inicio", 0)) for ev in comerciales
        )
        if t_primer_com is None:
            continue
        for i in idxs:
            ev = eventos_fase2_conductores[i]
            if (ev.get("evento") or "").strip() != "Parada":
                continue
            bus_ev = ev.get("bus")
            if bus_ev in (None, "", 0, "0"):
                continue
            ini_p = _tiempo_a_minutos(ev.get("inicio", 0))
            fin_p = _tiempo_a_minutos(ev.get("fin", 0))
            if ini_p is None or fin_p is None:
                continue
            # Parada totalmente antes del primer Comercial del conductor → eliminar.
            if fin_p <= t_primer_com:
                indices_a_eliminar_paradas_pre.add(i)
    if indices_a_eliminar_paradas_pre:
        eventos_fase2_conductores = [
            ev for idx, ev in enumerate(eventos_fase2_conductores) if idx not in indices_a_eliminar_paradas_pre
        ]

    # Ajustar FnS para que esté exactamente acoplado al último evento real
    # (Vacio, Comercial, Parada o Desplazamiento) del conductor, sin gap.
    eventos_por_cid_fns: Dict[Any, List[Dict[str, Any]]] = collections.defaultdict(list)
    for ev in eventos_fase2_conductores:
        cid = ev.get("conductor")
        if cid:
            eventos_por_cid_fns[cid].append(ev)
    for cid, lista in eventos_por_cid_fns.items():
        if not lista:
            continue
        # último evento real (no InS/FnS)
        reales = [e for e in lista if (e.get("evento") or "").strip() not in ("InS", "FnS")]
        if not reales:
            continue
        ultimo_real = max(reales, key=lambda e: _tiempo_a_minutos(e.get("fin", 0)))
        fin_real = _tiempo_a_minutos(ultimo_real.get("fin", 0))
        if fin_real is None:
            continue
        for ev in lista:
            if (ev.get("evento") or "").strip() != "FnS":
                continue
            ev["inicio"] = formatear_hora_deltatime(fin_real)
            ev["fin"] = formatear_hora_deltatime(fin_real)
            ev["duracion"] = "00:00"

    eventos_fase2_conductores.sort(
        key=lambda e: (
            _tiempo_a_minutos(e.get("inicio", 0)),
            str(e.get("conductor", "")),
            str(e.get("bus", "")),
            e.get("evento", ""),
        )
    )
    return eventos_fase2_conductores


def exportar_resultado_excel(
    config: Dict[str, Any],
    bloques_bus: List[List[Dict[str, Any]]],
    turnos_seleccionados: List[Dict[str, Any]],
    viajes_comerciales: List[Dict[str, Any]],
    metadata_tareas: Dict[Any, Dict[str, Any]],
    status_fase1: str,
    status_fase2: str,
    path_out: str = "resultado_diagramacion.xlsx",
    gestor: Optional[GestorDeLogistica] = None,
    verbose: bool = False,
    status_f3: Optional[str] = None,
    eventos_bus: Optional[List[List[Dict[str, Any]]]] = None,
) -> Optional[Dict[str, Any]]:
    """
    Exporta resultados a Excel (5 hojas), usando lógica de origen:
    eventos desde bloques, asignación conductor por minuto, InS/FnS/Desplazamiento.
    Returns:
        Dict con "conductores_exportados" o None si hubo error.
    """
    if gestor is None:
        gestor = GestorDeLogistica(config)
    path_out = str(Path(path_out).resolve())
    print(f"Iniciando exportación a {path_out}...")
    if not bloques_bus:
        print("ERROR: No hay bloques de buses para exportar.")
        return None
    if not turnos_seleccionados:
        print("ERROR: No hay turnos de conductores para exportar.")
        return None
    if not viajes_comerciales:
        print("ERROR: No hay viajes comerciales para exportar.")
        return None
    print(f"[EXPORTACIÓN] Bloques: {len(bloques_bus)} | Turnos: {len(turnos_seleccionados)} | Viajes: {len(viajes_comerciales)}")

    todos_eventos = _generar_eventos_completos(
        gestor, bloques_bus, turnos_seleccionados, viajes_comerciales, metadata_tareas, eventos_bus
    )
    conductores_con_comercial: Set[int] = set()
    for ev in todos_eventos:
        if str(ev.get("evento", "")).strip().upper() == "COMERCIAL" and ev.get("conductor"):
            try:
                conductores_con_comercial.add(int(ev["conductor"]))
            except (TypeError, ValueError):
                pass

    # Filtrar eventos de conductores que no tienen ningún viaje Comercial
    if conductores_con_comercial:
        eventos_filtrados_global: List[Dict[str, Any]] = []
        for ev in todos_eventos:
            cid = ev.get("conductor")
            if not cid:
                eventos_filtrados_global.append(ev)
                continue
            try:
                cid_int = int(cid)
            except (TypeError, ValueError):
                eventos_filtrados_global.append(ev)
                continue
            if cid_int in conductores_con_comercial:
                eventos_filtrados_global.append(ev)
        todos_eventos = eventos_filtrados_global

    # Renumerar conductores para que no queden huecos (ej. que el antiguo 122 pase a 121
    # si el 121 original no tenía viajes comerciales). Se preserva el orden por id.
    mapa_conductores: Dict[int, int] = {}
    if conductores_con_comercial:
        ids_ordenados = sorted(conductores_con_comercial)
        for nuevo_id, viejo_id in enumerate(ids_ordenados, start=1):
            mapa_conductores[viejo_id] = nuevo_id

        for ev in todos_eventos:
            cid = ev.get("conductor")
            if not cid:
                continue
            try:
                cid_int = int(cid)
            except (TypeError, ValueError):
                continue
            nuevo = mapa_conductores.get(cid_int)
            if nuevo is not None:
                ev["conductor"] = nuevo

    num_conductores_exportados = len(mapa_conductores) if conductores_con_comercial else 0

    wb = Workbook()
    font_bold = Font(bold=True)
    wrap_align = Alignment(wrap_text=True)

    # --- Hoja 0: ResumenOptimizacion ---
    ws0 = wb.active
    ws0.title = "ResumenOptimizacion"
    ws0.append(["Fase", "Estado Final", "Descripción"])
    ws0.append([
        "Fase 1: Buses",
        status_fase1,
        "OPTIMAL: Flota mínima garantizada." if status_fase1 == "OPTIMAL" else "FEASIBLE o otro.",
    ])
    ws0.append([
        "Fase 2: Conductores",
        status_fase2,
        "OPTIMAL: Número mínimo de conductores." if status_fase2 == "OPTIMAL" else "FEASIBLE o otro.",
    ])
    # Resumen Fase 3 coherente con los conductores realmente exportados
    if status_f3:
        desc_f3 = f"Fase 3: {len(turnos_seleccionados)} turnos -> {num_conductores_exportados} conductores exportados (solo con Comercial)."
        ws0.append(["Fase 3: Unión de Conductores", "COMPLETADA", desc_f3])
    ws0.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
    for cell in ws0[1]:
        cell.font = font_bold
    for col in ["A", "B", "C"]:
        ws0.column_dimensions[col].width = 25
    ws0["C2"].alignment = wrap_align
    ws0["C3"].alignment = wrap_align

    # --- Hoja 1: BloquesBuses ---
    ws1 = wb.create_sheet("BloquesBuses")
    ws1.append(["bus_id", "seq", "id_viaje", "linea", "sentido", "origen", "destino", "inicio_hhmm", "fin_hhmm", "duracion_min", "desc"])
    for b_id_idx, bloque in enumerate(bloques_bus):
        b_id = b_id_idx + 1
        for s, v in enumerate(bloque, start=1):
            dur = v.get("fin", 0) - v.get("inicio", 0)
            ws1.append([
                b_id, s, v.get("id", ""), v.get("linea", ""), v.get("sentido", ""),
                v.get("origen", ""), v.get("destino", ""),
                formatear_hora_deltatime(v.get("inicio", 0)), formatear_hora_deltatime(v.get("fin", 0)),
                dur, v.get("desc", ""),
            ])

    # --- Hoja 2: TurnosConductores ---
    ws2 = wb.create_sheet("TurnosConductores")
    ws2.append([
        "conductor_id", "bus_id_inicial", "inicio_jornada_hhmm", "fin_jornada_hhmm",
        "duracion_jornada_min", "num_tareas_comerciales", "tipo_mov_inicio", "tipo_mov_fin",
    ])
    for old_c_id, t in enumerate(turnos_seleccionados, start=1):
        # Solo exportar turnos cuyo conductor tenga al menos un Comercial (mapa_conductores)
        if conductores_con_comercial:
            try:
                nuevo_id = mapa_conductores.get(int(old_c_id))
            except (TypeError, ValueError):
                nuevo_id = None
            if not nuevo_id:
                continue
            c_id = nuevo_id
        else:
            c_id = old_c_id

        num_tareas = len(t.get("tareas_con_bus", []))
        bus_id_inicial = 0
        if t.get("tareas_con_bus"):
            bus_id_inicial = t["tareas_con_bus"][0][1] + 1
        ws2.append([
            c_id, bus_id_inicial,
            formatear_hora_deltatime(t.get("inicio", 0)), formatear_hora_deltatime(t.get("fin", 0)),
            t.get("duracion", 0), num_tareas,
            t.get("tipo_mov_inicio", "N/A"), t.get("tipo_mov_fin", "N/A"),
        ])

    # --- Hoja 3: BusEventos (Vacio, Parada, Comercial, Recarga por bus) ---
    eventos_por_bus: Dict[int, List[Dict]] = collections.defaultdict(list)
    for ev in todos_eventos:
        tipo_ev = (str(ev.get("evento", "")) or "").strip().upper()
        if tipo_ev not in ("VACIO", "PARADA", "COMERCIAL", "RECARGA"):
            continue
        bus_id = ev.get("bus")
        if bus_id is not None and bus_id != "":
            try:
                eventos_por_bus[int(bus_id)].append(ev)
            except (TypeError, ValueError):
                pass
    # Mapa bus_id -> tipo_bus desde bloques (para eventos sin tipo_bus)
    bus_tipo_map: Dict[int, str] = {}
    for bus_idx, bloque in enumerate(bloques_bus):
        bus_id = bus_idx + 1
        for viaje in bloque or []:
            tb = viaje.get("tipo_bus")
            if tb:
                bus_tipo_map[bus_id] = str(tb)
                break
    ws3 = wb.create_sheet("BusEventos")
    headers_bus_eventos = [
        "Evento", "Tipo", "Inicio", "De", "Fin", "A", "Duración", "Servicio", "Bus", "Tipo Bus",
        "Línea", "km", "V. Inferido", "Id.", "Sentido", "Tipo Mapeado", "Autonomía", "Consumo",
        "% Batería", "Pos. en P. Recarga", "Bus Orden",
    ]
    ws3.append(headers_bus_eventos)
    for bus_id in sorted(eventos_por_bus):
        eventos_ord = sorted(eventos_por_bus[bus_id], key=lambda e: _tiempo_a_minutos(e.get("inicio", 0)))
        for orden, ev in enumerate(eventos_ord, start=1):
            duracion_min = _tiempo_a_minutos(ev.get("duracion", 0)) or max(
                0, _tiempo_a_minutos(ev.get("fin", 0)) - _tiempo_a_minutos(ev.get("inicio", 0))
            )
            km = ev.get("kilometros", 0) or 0
            v_inferido = ""
            if duracion_min > 0 and km > 0:
                v_inferido = round(km / (duracion_min / 60.0), 1)
            tipo_bus = ev.get("tipo_bus") or bus_tipo_map.get(bus_id, "")
            autonomia = ""
            consumo = ""
            if tipo_bus and gestor and hasattr(gestor, "obtener_tipo_bus"):
                config_tipo = gestor.obtener_tipo_bus(tipo_bus)
                if config_tipo:
                    if config_tipo.autonomia_km is not None:
                        autonomia = config_tipo.autonomia_km
                    if config_tipo.parametros_electricos and km > 0:
                        cons_pct = config_tipo.parametros_electricos.consumo_pct_por_km
                        if cons_pct is not None:
                            consumo = round(cons_pct * km, 2)
            pct_bateria = ev.get("porcentaje_bateria", "") or ""
            pos_recarga = ev.get("posicion_recarga", "") or ""
            ws3.append([
                ev.get("evento", ""),
                ev.get("evento", ""),
                ev.get("inicio", ""),
                ev.get("origen", ""),
                ev.get("fin", ""),
                ev.get("destino", ""),
                ev.get("duracion", ""),
                ev.get("conductor", ""),
                bus_id,
                tipo_bus,
                ev.get("linea", ""),
                km,
                v_inferido,
                ev.get("viaje_id", ""),
                ev.get("sentido", ""),
                tipo_bus,
                autonomia,
                consumo,
                pct_bateria,
                pos_recarga,
                orden,
            ])

    # --- Hoja 4: EventosCompletos (columnas compatibles con analisis_completo.py) ---
    ws4 = wb.create_sheet("EventosCompletos")
    ws4.append([
        "Tipo", "Bus", "Conductor", "Inicio", "Fin", "Duración",
        "Origen", "Destino", "km", "Línea", "desc",
    ])
    for ev in todos_eventos:
        ws4.append([
            ev.get("evento", ""), ev.get("bus", ""), ev.get("conductor", ""),
            ev.get("inicio", ""), ev.get("fin", ""), ev.get("duracion", ""),
            ev.get("origen", ""), ev.get("destino", ""), ev.get("kilometros", 0),
            ev.get("linea", ""), ev.get("desc", ""),
        ])

    try:
        wb.save(path_out)
        print(f"[OK] Archivo exportado correctamente: {path_out}")
    except PermissionError:
        import os
        path_eval = os.path.join(os.path.dirname(path_out), "resultado_diagramacion_eval.xlsx")
        try:
            wb.save(path_eval)
            print(f"Permiso denegado en archivo principal. Exportado a: {path_eval}")
        except Exception:
            print(f"Error: Permiso denegado. Cierra '{path_out}'.")
            return None
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        return None

    resumen = collections.Counter(ev.get("evento") for ev in todos_eventos)
    print("Resumen de eventos exportados:")
    for tipo, cnt in sorted(resumen.items(), key=lambda x: -x[1]):
        print(f"  - {tipo}: {cnt}")
    return {"conductores_exportados": num_conductores_exportados}
