"""
Análisis exhaustivo de continuidad:
1. Continuidad de nodos: destino(evento N) == origen(evento N+1)
2. Continuidad temporal: inicio(evento N+1) == fin(evento N)
3. Por conductor y por bus
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
from collections import defaultdict

def cargar_configuracion() -> Dict[str, Any]:
    """Carga la configuración"""
    config_path = Path(__file__).parent / "configuracion.json"
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def normalizar_nodo(nodo: str) -> str:
    """Normaliza el nombre del nodo"""
    if not nodo:
        return ""
    return str(nodo).strip().upper()

def mismo_nodo(nodo1: str, nodo2: str) -> bool:
    """Verifica si dos nodos son el mismo"""
    n1 = normalizar_nodo(nodo1)
    n2 = normalizar_nodo(nodo2)
    if n1 == n2:
        return True
    # Variantes comunes
    n1_sin = n1.replace("DEPOSITO", "").replace("DEPÓSITO", "").strip()
    n2_sin = n2.replace("DEPOSITO", "").replace("DEPÓSITO", "").strip()
    if n1_sin and n2_sin and n1_sin == n2_sin:
        return True
    return False

def analizar_continuidad_excel(archivo_excel: Optional[str] = None) -> Dict[str, Any]:
    """Analiza la continuidad en el archivo Excel generado"""
    import openpyxl
    
    if archivo_excel is None:
        archivo_excel = Path(__file__).parent / "resultado_diagramacion.xlsx"
    
    if not Path(archivo_excel).exists():
        return {"error": f"Archivo no encontrado: {archivo_excel}"}
    
    wb = openpyxl.load_workbook(archivo_excel)
    
    # Buscar hoja de eventos (puede llamarse "EventosCompletos" o similar)
    nombre_hoja_eventos = None
    for nombre_hoja in wb.sheetnames:
        nombre_lower = nombre_hoja.lower()
        if "evento" in nombre_lower or "completo" in nombre_lower:
            nombre_hoja_eventos = nombre_hoja
            break
    
    # Si no se encuentra, usar la última hoja (generalmente es la de eventos)
    if not nombre_hoja_eventos and wb.sheetnames:
        nombre_hoja_eventos = wb.sheetnames[-1]
    
    if not nombre_hoja_eventos:
        return {"error": f"Hoja de eventos no encontrada. Hojas: {', '.join(wb.sheetnames)}"}
    
    ws = wb[nombre_hoja_eventos]
    
    # Leer encabezados (buscar por posición conocida o por nombre)
    headers = {}
    header_row = ws[1]
    for col_idx, cell in enumerate(header_row, start=1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    # Mapeo de nombres alternativos
    nombre_conductor = None
    nombre_bus = None
    nombre_evento = None
    nombre_origen = None
    nombre_destino = None
    nombre_inicio = None
    nombre_fin = None
    
    for header, col_idx in headers.items():
        header_lower = header.lower()
        if "conductor" in header_lower or "servicio" in header_lower:
            nombre_conductor = header
        if "bus" in header_lower:
            nombre_bus = header
        if "evento" in header_lower:
            nombre_evento = header
        if "origen" in header_lower:
            nombre_origen = header
        if "destino" in header_lower:
            nombre_destino = header
        if "inicio" in header_lower:
            nombre_inicio = header
        if "fin" in header_lower:
            nombre_fin = header
    
    # Leer eventos
    eventos = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        evento = {}
        for header, col_idx in headers.items():
            if col_idx <= len(row):
                evento[header] = row[col_idx - 1].value
        
        # Normalizar nombres de campos
        if nombre_conductor:
            evento["Conductor"] = evento.get(nombre_conductor)
        if nombre_bus:
            evento["Bus"] = evento.get(nombre_bus)
        if nombre_evento:
            evento["Evento"] = evento.get(nombre_evento)
        if nombre_origen:
            evento["Origen"] = evento.get(nombre_origen)
        if nombre_destino:
            evento["Destino"] = evento.get(nombre_destino)
        if nombre_inicio:
            evento["Inicio"] = evento.get(nombre_inicio)
        if nombre_fin:
            evento["Fin"] = evento.get(nombre_fin)
        
        # Convertir valores de tiempo (pueden venir como "05:30" o como minutos)
        def convertir_tiempo_a_minutos(valor):
            if valor is None:
                return 0
            if isinstance(valor, (int, float)):
                return float(valor)
            valor_str = str(valor).strip()
            if ":" in valor_str:
                # Formato hora "HH:MM"
                partes = valor_str.split(":")
                if len(partes) >= 2:
                    try:
                        horas = int(partes[0])
                        minutos = int(partes[1])
                        return horas * 60 + minutos
                    except (ValueError, TypeError):
                        pass
            try:
                return float(valor_str)
            except (ValueError, TypeError):
                return 0
        
        inicio = convertir_tiempo_a_minutos(evento.get("Inicio", 0))
        fin = convertir_tiempo_a_minutos(evento.get("Fin", 0))
        
        evento["_inicio"] = inicio
        evento["_fin"] = fin
        evento["_row"] = row_idx
        
        if evento.get("Evento"):
            eventos.append(evento)
    
    # Analizar por conductor
    errores_conductor = []
    por_conductor = defaultdict(list)
    
    for ev in eventos:
        conductor = ev.get("Conductor")
        # El campo puede estar vacío, ser un número, o un string
        if conductor is not None:
            try:
                if isinstance(conductor, (int, float)):
                    cid = int(conductor)
                elif isinstance(conductor, str):
                    cid_val = conductor.strip()
                    if cid_val and cid_val.lower() not in ["", "none", "null"]:
                        cid = int(float(cid_val))
                    else:
                        cid = None
                else:
                    cid_val = str(conductor).strip()
                    if cid_val and cid_val.lower() not in ["", "none", "null"]:
                        cid = int(float(cid_val))
                    else:
                        cid = None
                if cid is not None and cid > 0:
                    por_conductor[cid].append(ev)
            except (ValueError, TypeError):
                pass
    
    for cid, evs_conductor in por_conductor.items():
        evs_ord = sorted(evs_conductor, key=lambda e: (e.get("_inicio", 0), e.get("_fin", 0)))
        
        for i in range(1, len(evs_ord)):
            ant = evs_ord[i - 1]
            act = evs_ord[i]
            
            dest_ant = str(ant.get("Destino", "")).strip()
            orig_act = str(act.get("Origen", "")).strip()
            fin_ant = ant.get("_fin", 0)
            ini_act = act.get("_inicio", 0)
            
            # Verificar continuidad de nodos
            if not mismo_nodo(dest_ant, orig_act):
                errores_conductor.append({
                    "tipo": "nodo",
                    "conductor": cid,
                    "fila_ant": ant.get("_row"),
                    "fila_act": act.get("_row"),
                    "evento_ant": ant.get("Evento"),
                    "evento_act": act.get("Evento"),
                    "destino_anterior": dest_ant,
                    "origen_actual": orig_act,
                    "mensaje": f"Conductor {cid}: nodo fin evento anterior ({dest_ant}) != nodo inicio siguiente ({orig_act})"
                })
            
            # Verificar continuidad temporal
            if abs(fin_ant - ini_act) > 1:  # Tolerancia de 1 minuto
                errores_conductor.append({
                    "tipo": "tiempo",
                    "conductor": cid,
                    "fila_ant": ant.get("_row"),
                    "fila_act": act.get("_row"),
                    "evento_ant": ant.get("Evento"),
                    "evento_act": act.get("Evento"),
                    "fin_anterior": fin_ant,
                    "inicio_actual": ini_act,
                    "gap": ini_act - fin_ant,
                    "mensaje": f"Conductor {cid}: fin evento anterior ({fin_ant}) != inicio siguiente ({ini_act}), gap: {ini_act - fin_ant} min"
                })
    
    # Analizar por bus
    errores_bus = []
    por_bus = defaultdict(list)
    
    for ev in eventos:
        bus = ev.get("Bus")
        if bus and str(bus).strip():
            try:
                bus_id = int(bus)
                por_bus[bus_id].append(ev)
            except (ValueError, TypeError):
                pass
    
    for bus_id, evs_bus in por_bus.items():
        evs_ord = sorted(evs_bus, key=lambda e: (e.get("_inicio", 0), e.get("_fin", 0)))
        
        for i in range(1, len(evs_ord)):
            ant = evs_ord[i - 1]
            act = evs_ord[i]
            
            dest_ant = str(ant.get("Destino", "")).strip()
            orig_act = str(act.get("Origen", "")).strip()
            fin_ant = ant.get("_fin", 0)
            ini_act = act.get("_inicio", 0)
            
            # Verificar continuidad de nodos
            if not mismo_nodo(dest_ant, orig_act):
                errores_bus.append({
                    "tipo": "nodo",
                    "bus": bus_id,
                    "fila_ant": ant.get("_row"),
                    "fila_act": act.get("_row"),
                    "evento_ant": ant.get("Evento"),
                    "evento_act": act.get("Evento"),
                    "destino_anterior": dest_ant,
                    "origen_actual": orig_act,
                    "mensaje": f"Bus {bus_id}: nodo fin evento anterior ({dest_ant}) != nodo inicio siguiente ({orig_act})"
                })
            
            # Verificar continuidad temporal
            if abs(fin_ant - ini_act) > 1:  # Tolerancia de 1 minuto
                errores_bus.append({
                    "tipo": "tiempo",
                    "bus": bus_id,
                    "fila_ant": ant.get("_row"),
                    "fila_act": act.get("_row"),
                    "evento_ant": ant.get("Evento"),
                    "evento_act": act.get("Evento"),
                    "fin_anterior": fin_ant,
                    "inicio_actual": ini_act,
                    "gap": ini_act - fin_ant,
                    "mensaje": f"Bus {bus_id}: fin evento anterior ({fin_ant}) != inicio siguiente ({ini_act}), gap: {ini_act - fin_ant} min"
                })
    
    return {
        "errores_conductor": errores_conductor,
        "errores_bus": errores_bus,
        "total_eventos": len(eventos),
        "total_conductores": len(por_conductor),
        "total_buses": len(por_bus),
    }

def main():
    """Función principal"""
    print("=" * 80)
    print("ANÁLISIS DE CONTINUIDAD COMPLETA")
    print("=" * 80)
    print()
    
    resultado = analizar_continuidad_excel()
    
    if "error" in resultado:
        print(f"[ERROR] {resultado['error']}")
        return
    
    errores_conductor = resultado["errores_conductor"]
    errores_bus = resultado["errores_bus"]
    
    print(f"Total eventos analizados: {resultado['total_eventos']}")
    print(f"Total conductores: {resultado['total_conductores']}")
    print(f"Total buses: {resultado['total_buses']}")
    print()
    
    total_errores = len(errores_conductor) + len(errores_bus)
    
    if total_errores == 0:
        print("[OK] Continuidad perfecta: todos los eventos están conectados correctamente")
        print()
        return
    
    print(f"[ERROR] Se encontraron {total_errores} errores de continuidad:")
    print()
    
    if errores_conductor:
        print(f"Errores por CONDUCTOR ({len(errores_conductor)}):")
        errores_nodo_c = [e for e in errores_conductor if e["tipo"] == "nodo"]
        errores_tiempo_c = [e for e in errores_conductor if e["tipo"] == "tiempo"]
        
        if errores_nodo_c:
            print(f"  - Desconexiones de nodos: {len(errores_nodo_c)}")
            for error in errores_nodo_c[:10]:
                print(f"    {error['mensaje']} (Fila {error['fila_ant']} -> {error['fila_act']})")
            if len(errores_nodo_c) > 10:
                print(f"    ... y {len(errores_nodo_c) - 10} más")
        
        if errores_tiempo_c:
            print(f"  - Desconexiones temporales: {len(errores_tiempo_c)}")
            for error in errores_tiempo_c[:10]:
                print(f"    {error['mensaje']} (Fila {error['fila_ant']} -> {error['fila_act']})")
            if len(errores_tiempo_c) > 10:
                print(f"    ... y {len(errores_tiempo_c) - 10} más")
        print()
    
    if errores_bus:
        print(f"Errores por BUS ({len(errores_bus)}):")
        errores_nodo_b = [e for e in errores_bus if e["tipo"] == "nodo"]
        errores_tiempo_b = [e for e in errores_bus if e["tipo"] == "tiempo"]
        
        if errores_nodo_b:
            print(f"  - Desconexiones de nodos: {len(errores_nodo_b)}")
            for error in errores_nodo_b[:10]:
                print(f"    {error['mensaje']} (Fila {error['fila_ant']} -> {error['fila_act']})")
            if len(errores_nodo_b) > 10:
                print(f"    ... y {len(errores_nodo_b) - 10} más")
        
        if errores_tiempo_b:
            print(f"  - Desconexiones temporales: {len(errores_tiempo_b)}")
            for error in errores_tiempo_b[:10]:
                print(f"    {error['mensaje']} (Fila {error['fila_ant']} -> {error['fila_act']})")
            if len(errores_tiempo_b) > 10:
                print(f"    ... y {len(errores_tiempo_b) - 10} más")
        print()

if __name__ == "__main__":
    main()
