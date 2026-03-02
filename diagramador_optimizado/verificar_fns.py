"""Script para verificar FnS: duración 0 y eventos después del FnS"""
import openpyxl
from pathlib import Path

archivo = Path(__file__).parent / "resultado_diagramacion.xlsx"
if not archivo.exists():
    print(f"Archivo no encontrado: {archivo}")
    exit(1)

wb = openpyxl.load_workbook(archivo)
ws = wb['EventosCompletos']

# Leer todos los eventos
eventos = []
for i in range(2, ws.max_row + 1):
    evento = ws.cell(i, 1).value
    conductor = ws.cell(i, 3).value
    inicio = ws.cell(i, 4).value
    fin = ws.cell(i, 5).value
    duracion = ws.cell(i, 6).value
    eventos.append({
        'fila': i,
        'evento': str(evento).strip() if evento else '',
        'conductor': conductor,
        'inicio': inicio,
        'fin': fin,
        'duracion': duracion
    })

# Convertir tiempos a minutos para comparación
def tiempo_a_minutos(t):
    if t is None:
        return 0
    if isinstance(t, (int, float)):
        return int(t)
    if isinstance(t, str) and ':' in t:
        partes = t.split(':')
        if len(partes) >= 2:
            try:
                horas = int(partes[0])
                minutos = int(partes[1])
                return horas * 60 + minutos
            except:
                pass
    return 0

# Agrupar por conductor
por_conductor = {}
for ev in eventos:
    cid = ev['conductor']
    if cid is None:
        continue
    try:
        cid = int(cid)
    except:
        continue
    if cid not in por_conductor:
        por_conductor[cid] = []
    por_conductor[cid].append(ev)

print("=" * 80)
print("VERIFICACIÓN DE FnS: DURACIÓN Y EVENTOS DESPUÉS")
print("=" * 80)
print()

errores_fns_duracion = []
errores_eventos_despues = []

for cid in sorted(por_conductor.keys()):
    evs = sorted(por_conductor[cid], key=lambda e: (tiempo_a_minutos(e['inicio']), tiempo_a_minutos(e['fin'])))
    
    # Buscar FnS
    fns = next((e for e in evs if str(e['evento']).strip().upper() == 'FNS'), None)
    if not fns:
        continue
    
    inicio_fns = tiempo_a_minutos(fns['inicio'])
    fin_fns = tiempo_a_minutos(fns['fin'])
    duracion_fns = fns['duracion']
    
    # Verificar duración
    if inicio_fns != fin_fns or (duracion_fns is not None and duracion_fns != 0):
        errores_fns_duracion.append({
            'conductor': cid,
            'fila': fns['fila'],
            'inicio': fns['inicio'],
            'fin': fns['fin'],
            'duracion': duracion_fns,
            'duracion_calculada': fin_fns - inicio_fns
        })
    
    # Buscar eventos después del FnS
    eventos_despues = [e for e in evs if e is not fns and tiempo_a_minutos(e['inicio']) >= fin_fns]
    if eventos_despues:
        errores_eventos_despues.append({
            'conductor': cid,
            'fns': fns,
            'eventos_despues': eventos_despues
        })

if errores_fns_duracion:
    print(f"[ERROR] FnS con duración incorrecta: {len(errores_fns_duracion)}")
    for err in errores_fns_duracion[:10]:
        print(f"  - Conductor {err['conductor']} (Fila {err['fila']}): Inicio={err['inicio']}, Fin={err['fin']}, "
              f"Duración={err['duracion']}, Duración calculada={err['duracion_calculada']}")
    if len(errores_fns_duracion) > 10:
        print(f"  ... y {len(errores_fns_duracion) - 10} más")
    print()
else:
    print("[OK] Todos los FnS tienen duración 0 (inicio == fin)")
    print()

if errores_eventos_despues:
    print(f"[ERROR] Eventos después del FnS: {len(errores_eventos_despues)} conductores")
    for err in errores_eventos_despues[:10]:
        cid = err['conductor']
        fns = err['fns']
        eventos = err['eventos_despues']
        print(f"  - Conductor {cid}: FnS en fila {fns['fila']} (fin={fns['fin']}), "
              f"{len(eventos)} eventos después:")
        for ev in eventos[:3]:
            print(f"      Fila {ev['fila']}: {ev['evento']}, inicio={ev['inicio']}")
        if len(eventos) > 3:
            print(f"      ... y {len(eventos) - 3} más")
    if len(errores_eventos_despues) > 10:
        print(f"  ... y {len(errores_eventos_despues) - 10} conductores más")
    print()
else:
    print("[OK] No hay eventos después del FnS para ningún conductor")
    print()

# Verificar conductor 1 específicamente
if 1 in por_conductor:
    evs_1 = sorted(por_conductor[1], key=lambda e: (tiempo_a_minutos(e['inicio']), tiempo_a_minutos(e['fin'])))
    fns_1 = next((e for e in evs_1 if str(e['evento']).strip().upper() == 'FNS'), None)
    if fns_1:
        print(f"Conductor 1 - FnS:")
        print(f"  Fila: {fns_1['fila']}")
        print(f"  Inicio: {fns_1['inicio']}")
        print(f"  Fin: {fns_1['fin']}")
        print(f"  Duración: {fns_1['duracion']}")
        fin_fns_1 = tiempo_a_minutos(fns_1['fin'])
        eventos_despues_1 = [e for e in evs_1 if e is not fns_1 and tiempo_a_minutos(e['inicio']) >= fin_fns_1]
        print(f"  Eventos después del FnS: {len(eventos_despues_1)}")
        if eventos_despues_1:
            for ev in eventos_despues_1[:5]:
                print(f"    - Fila {ev['fila']}: {ev['evento']}, inicio={ev['inicio']}, fin={ev['fin']}")

print()
print("=" * 80)
total_errores = len(errores_fns_duracion) + len(errores_eventos_despues)
if total_errores == 0:
    print("[OK] Todas las validaciones pasaron correctamente")
else:
    print(f"[ERROR] Total de errores encontrados: {total_errores}")
